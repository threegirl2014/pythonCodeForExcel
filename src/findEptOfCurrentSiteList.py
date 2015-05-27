#-*- coding:utf-8 -*





from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


import os
import time

def findSiteId(rawSiteId):
	length = len(rawSiteId)
	if length == 4:
		siteId = rawSiteId
	elif length > 4:
		siteId = rawSiteId[-4:]
	elif length < 4:
		siteId = '0'*(4-length) + rawSiteId
	return siteId

def calculateTrxgroup(trxNum, lastCharacter, sectorScenario):
	trxgroup = ''
	sectorEqmNum = ''
	lastCharacterNum = ''
	if( lastCharacter == 'A' ):
		lastCharacterNum = '1'
		if( sectorScenario == 13 ):
			sectorEqmNum = '2'
			if( trxNum <= 8 ):
				trxgroup += ('1' + ',1' * 2  + ',4' * (trxNum - 3) )
			elif( trxNum <= 11 ):
				trxgroup += ('1' + ',1' * 3 + ',4' * (trxNum - 4) )
			elif( trxNum >= 12 ):
				trxgroup += ('1' + ',1' * 4 + ',4' * (trxNum - 5) )
		elif( sectorScenario == 14 or sectorScenario == 16 or sectorScenario == 29 ):
			sectorEqmNum = '1'
			trxgroup += '1'
	elif( lastCharacter == 'B' ):
		lastCharacterNum = '2'
		if( sectorScenario == 13 ):
			sectorEqmNum = '2'
			if( trxNum <= 8 ):
				trxgroup += ('2' + ',2' * 2  + ',5' * (trxNum - 3) )
			elif( trxNum <= 11 ):
				trxgroup += ('2' + ',2' * 3 + ',5' * (trxNum - 4) )
			elif( trxNum >= 12 ):
				trxgroup += ('2' + ',2' * 4 + ',5' * (trxNum - 5) )
		elif( sectorScenario == 14 or sectorScenario == 16 or sectorScenario == 29 ):
			sectorEqmNum = '1'
			trxgroup += '2'
	elif( lastCharacter == 'C' ):
		lastCharacterNum = '3'
		if( sectorScenario == 13 ):
			sectorEqmNum = '2'
			if( trxNum <= 8 ):
				trxgroup += ('3' + ',3' * 2  + ',6' * (trxNum - 3) )
			elif( trxNum <= 11 ):
				trxgroup += ('3' + ',3' * 3 + ',6' * (trxNum - 4) )
			elif( trxNum >= 12 ):
				trxgroup += ('3' + ',3' * 4 + ',6' * (trxNum - 5) )
		elif( sectorScenario == 14 or sectorScenario == 16 or sectorScenario == 29 ):
			sectorEqmNum = '1'
			trxgroup += '3'

	return [trxgroup, sectorEqmNum, lastCharacterNum]



currentSiteWb = load_workbook( filename = r'input\\result.xlsx' )
print "Worksheet name(s):", currentSiteWb.get_sheet_names()

currentSiteWs = currentSiteWb['result']

siteIdList = []

print currentSiteWs.get_highest_row()
for index in range(2, currentSiteWs.get_highest_row() + 1):
	siteIdList.append( findSiteId( str(currentSiteWs.cell( row = index, column = 1 ).value) ) )
	print str(currentSiteWs.cell( row = index, column = 1 ).value)

rruWb = load_workbook( filename = r'input\\rru info.xlsx')
print "Worksheet name(s): ", rruWb.get_sheet_names()

rruWs = rruWb['rru info']
print "Worksheet Title: ", rruWs.title

rruInfoDic = {}

for index in range( 2, rruWs.get_highest_row() + 1 ):
	siteId = findSiteId( str(rruWs.cell( row = index, column = 1).value) )
	sector = [
	rruWs.cell( row = index, column = 2 ).value, 
	rruWs.cell( row = index, column = 3 ).value, 
	rruWs.cell( row = index, column = 4 ).value 
	]
	rruInfoDic[siteId] = sector
print rruInfoDic

eptWb = load_workbook( filename = r'input\\EPT_GU.xlsx')
print "Worksheet name(s): ", eptWb.get_sheet_names()

umtsCellWs = eptWb['UMTS']
gsmCellWs = eptWb['GSM']
print "Worksheet Title: ", umtsCellWs.title
print "Worksheet Title: ", gsmCellWs.title

#for gsm
bscIdList = []
gCellSiteIdList = []
gCellNameList = []
gCellMccList = []
gCellMncList = []
gCellLacList = []
gCellIdList = []
gCellNccList = []
gCellBccList = []
gCellPowerList = []
gCellBcchList = []
gCellHsnList = []
gCellLatitudeList = []
gCellLatitudeIntList = []
gCellLatitudeDecimalList = []
gCellLongitudeList = []
gCellLongitudeIntList = []
gCellLongitudeDecimalList = []
gCellAzimuthList = []
gCellAltitudeIntList = []
gCellFreListList = []
gCellFreMaGroupListList = []
gCellPdchNumList = []
gCellSdcchNumList = []

bscOpcList = []

gCellTrxNumList = []
gCellTrxGroupList = []
gCellSectorEqmNumList = []
gCellLastCharacterList = []

# for umts
rncIdList = []
logicalRncIdList = []
uCellSiteIdList = []
# nodebList = []
uCellIdList = []
uCellNameList = []
uCellLacList = []
uCellSacList = []
uCellRacList = []
uCellUarfcnList = []
uCellPscList = []
uCellMaxDlPowerList = []
uCellPcpichList = []
uCellTcellList = []
uCellLatitudeList = []
uCellLongitudeList = []
uCellAzimuthList = []
uCellHeightList = []
uCellElectricalTiltList = []
uCellMechanicalTiltList = []
uCellSectorIdList = []

uCellSectorEqmIdList = []
uCellBasebandEqmIdList = []
uCellUplinkUarfcnList = []
uCellBandIndicatorList = []
uCellFrequencyBandwidthList = []



#need modify?
uCellUra = 10
maxCoverage = 5000
antennaOpening = 65

siteIdDic = {}
for index in range( 1,len(siteIdList) + 1 ):
	siteIdDic[siteIdList[index - 1]] = 1

for siteId, isExist in siteIdDic.items():
	print "%s, %d" % (siteId, isExist)

for index in range( 2, gsmCellWs.get_highest_row() + 1 ):
	temp = str(gsmCellWs.cell( row = index, column = 1 ).value)
	currentSiteId = findSiteId(temp)
	if( siteIdDic.has_key( currentSiteId ) ):
		print currentSiteId
		bscIdList.append( gsmCellWs.cell( row = index, column = 2 ).value )
		
		gCellSiteIdList.append( currentSiteId )

		gCellName = gsmCellWs.cell( row = index, column = 3 ).value
		gCellNameList.append( gCellName )
		gCellMccList.append( gsmCellWs.cell( row = index, column = 4 ).value )
		gCellMncList.append( gsmCellWs.cell( row = index, column = 5 ).value )
		gCellLacList.append( gsmCellWs.cell( row = index, column = 6 ).value )
		gCellIdList.append( gsmCellWs.cell( row = index, column = 7 ).value )
		gCellNccList.append( gsmCellWs.cell( row = index, column = 8 ).value )
		gCellBccList.append( gsmCellWs.cell( row = index, column = 9 ).value )

		if gsmCellWs.cell( row = index, column = 10 ).value and gsmCellWs.cell( row = index, column = 10 ).value != '#N/A':
			gCellPowerList.append( int(gsmCellWs.cell( row = index, column = 10 ).value * 10.0) )
		else:
			gCellPowerList.append( -999 )

		gCellBcchList.append( gsmCellWs.cell( row = index, column = 11 ).value )
		gCellHsnList.append( gsmCellWs.cell( row = index, column = 12 ).value )
	
		if gsmCellWs.cell( row = index, column = 13 ).value and gsmCellWs.cell( row = index, column = 13 ).value != '#N/A':
			latitude = gsmCellWs.cell( row = index, column = 13 ).value
			gCellLatitudeList.append( latitude )
			gCellLatitudeIntList.append( int(latitude) )
			gCellLatitudeDecimalList.append( int((latitude - int(latitude)) * 100000) )
		else:
			gCellLatitudeList.append( -999 )
			gCellLatitudeIntList.append( -999 )
			gCellLatitudeDecimalList.append( -999 )

		if gsmCellWs.cell( row = index, column = 14 ).value and gsmCellWs.cell( row = index, column = 14 ).value != '#N/A':
			longitude = gsmCellWs.cell( row = index, column = 14 ).value
			gCellLongitudeList.append( longitude )
			gCellLongitudeIntList.append( abs(int(longitude)) )
			gCellLongitudeDecimalList.append( abs(int((longitude - int(longitude)) * 100000)) )
		else:
			gCellLongitudeList.append( -999 )
			gCellLongitudeIntList.append( -999 )
			gCellLongitudeDecimalList.append( -999 )

		gCellAzimuthList.append( gsmCellWs.cell( row = index, column = 15 ).value )
		if gsmCellWs.cell( row = index, column = 16 ).value and gsmCellWs.cell( row = index, column = 16 ).value != '#N/A':
			gCellAltitudeIntList.append( int(gsmCellWs.cell( row = index, column = 16 ).value) )
		else:
			gCellAltitudeIntList.append( -999 )

		trxNum = gsmCellWs.cell( row = index, column = 18 ).value
		gCellTrxNumList.append( gsmCellWs.cell( row = index, column = 18 ).value )

		freList = str(gsmCellWs.cell( row = index, column = 17 ).value)
		#remove the space at the string end
		freListAfterStrip = freList.strip()
		#replace teh space between numbers with comma
		freListAfterRaplace = freListAfterStrip.replace(' ',',')
		freListAfterSplit = freListAfterRaplace.split(',')
		print trxNum, freListAfterSplit
		fre = ''
		for i in range(0,trxNum - 1):
			fre = fre + str(freListAfterSplit[i]) + ','
		gCellFreListList.append( fre[:-1] )
		freMaGroup = '[' + freListAfterRaplace + ']'
		gCellFreMaGroupListList.append( freMaGroup )

		# trxNum = freListAfterRaplace.count(',') + 2


		lastCharacter = gCellName[6]
		print currentSiteId, lastCharacter
		if lastCharacter == 'A' or lastCharacter == 'B' or lastCharacter == 'C':
			sectorScenario = rruInfoDic[currentSiteId][ord(lastCharacter) - 65]
		else:
			sectorScenario = -999
		print currentSiteId, trxNum, lastCharacter, sectorScenario
		afterCalculate = calculateTrxgroup(trxNum, lastCharacter, sectorScenario)
		gCellTrxGroupList.append( afterCalculate[0] )
		gCellSectorEqmNumList.append( afterCalculate[1] )
		gCellLastCharacterList.append( afterCalculate[2] )

		gCellSdcchNumList.append( gsmCellWs.cell( row = index, column = 19 ).value )
		gCellPdchNumList.append( gsmCellWs.cell( row = index, column = 20 ).value )

		bscOpcList.append( gsmCellWs.cell( row = index, column = 21 ).value )
		# gCellNseiQueueList.append( gsmCellWs.cell( row = index, column = 21 ).value )
		# gcellBvciBeginList.append( gsmCellWs.cell( row = index, column = 22 ).value )
		



for index in range( 2, umtsCellWs.get_highest_row() + 1 ):
	if not umtsCellWs.cell(row = index, column = 1).value:
		continue
	else:
		temp = str(umtsCellWs.cell(row = index, column = 1).value)
		siteIdWhole = findSiteId(temp)
		if( siteIdDic.has_key( siteIdWhole ) ):
			rncIdList.append( umtsCellWs.cell( row = index, column = 2 ).value )
			logicalRncIdList.append( umtsCellWs.cell( row = index, column = 20 ).value )
			print siteIdWhole
			uCellSiteIdList.append( siteIdWhole )
			uCellIdList.append( umtsCellWs.cell( row = index, column = 3 ).value )
			uCellNameList.append( str(umtsCellWs.cell( row = index, column = 4 ).value) )
			uCellLacList.append( umtsCellWs.cell( row = index, column = 5 ).value )
			uCellSacList.append( umtsCellWs.cell( row = index, column = 6 ).value )
			uCellRacList.append( umtsCellWs.cell( row = index, column = 7 ).value )
			uCellPscList.append( umtsCellWs.cell( row = index, column = 8 ).value )
			uCellMaxDlPowerList.append( umtsCellWs.cell( row = index, column = 9 ).value )
			uCellPcpichList.append( umtsCellWs.cell( row = index, column = 10 ).value )
			tcellLower = umtsCellWs.cell( row = index, column = 11 ).value
			if tcellLower:
				tcellUpper = tcellLower[:4].upper() + tcellLower[4:]
			uCellTcellList.append( tcellUpper )
			if umtsCellWs.cell( row = index, column = 12 ).value and umtsCellWs.cell( row = index, column = 12 ).value != '#N/A':
				uCellLatitudeList.append( int(1000000 * umtsCellWs.cell( row = index, column = 12 ).value) )
			else:
				uCellLatitudeList.append( -999 )
			if umtsCellWs.cell( row = index, column = 13 ).value and umtsCellWs.cell( row = index, column = 13 ).value != '#N/A':
				uCellLongitudeList.append( int(1000000 * umtsCellWs.cell( row = index, column = 13 ).value) )
			else:
				uCellLongitudeList.append( -999 )
			uCellAzimuthList.append( umtsCellWs.cell( row = index, column = 14 ).value )
			uCellHeightList.append( umtsCellWs.cell( row = index, column = 15 ).value )

			if umtsCellWs.cell( row = index, column = 16).value and umtsCellWs.cell( row = index, column = 16).value != '#N/A':
				uCellElectricalTiltList.append( int(umtsCellWs.cell( row = index, column = 16).value) * 10 )
			else:
				uCellElectricalTiltList.append( -999 )
			if umtsCellWs.cell( row = index, column = 17).value and umtsCellWs.cell( row = index, column = 17).value != '#N/A':
				uCellMechanicalTiltList.append( int(umtsCellWs.cell( row = index, column = 17).value) * 10 )
			else:
				uCellMechanicalTiltList.append( -999 )

			sectorId = umtsCellWs.cell(row = index, column = 18).value
			uCellSectorIdList.append( sectorId )

			uarfcn = str(umtsCellWs.cell(row = index, column = 19).value)
			uCellUarfcnList.append( uarfcn )

			sectorScenario = rruInfoDic[siteIdWhole][sectorId - 1]


			if( uarfcn == '9715' ):
				uCellBasebandEqmIdList.append(1)
				uCellUplinkUarfcnList.append(9315)
				uCellBandIndicatorList.append('Band2')
				uCellFrequencyBandwidthList.append(4200)
				if( sectorScenario == 13 or sectorScenario == 14 ):
					if( sectorId == 1 ):
						uCellSectorEqmIdList.append(7)
					elif( sectorId == 2 ):
						uCellSectorEqmIdList.append(8)
					elif( sectorId == 3 ):
						uCellSectorEqmIdList.append(9)
					else:
						uCellSectorEqmIdList.append(0)
				elif( sectorScenario == 15 or sectorScenario == 29  ):
					if( sectorId == 1 ):
						uCellSectorEqmIdList.append(16)
					elif( sectorId == 2 ):
						uCellSectorEqmIdList.append(17)
					elif( sectorId == 3 ):
						uCellSectorEqmIdList.append(18)
				else:
					uCellSectorEqmIdList.append(-999 )					
			else:
				uCellBasebandEqmIdList.append(0)
				uCellUplinkUarfcnList.append( int(uarfcn) - 225 )
				uCellBandIndicatorList.append('Band5')
				uCellFrequencyBandwidthList.append(5000)
				if( sectorScenario == 13 or sectorScenario == 14 or sectorScenario == 15 or 
					sectorScenario == 16 or sectorScenario == 24 or sectorScenario == 29 ):
					if( sectorId == 1 ):
						uCellSectorEqmIdList.append(4)
					elif( sectorId == 2 ):
						uCellSectorEqmIdList.append(5)
					elif( sectorId == 3 ):
						uCellSectorEqmIdList.append(6)
				else:
					uCellSectorEqmIdList.append( -999 )


eptCellInfoWb = Workbook()
gsmCellConfigurationWs = eptCellInfoWb.create_sheet(0,'GSM Cell')
umtsCellConfigurationWs = eptCellInfoWb.create_sheet(1, 'UMTS Cell')

gsmCellConfigurationWs.append(
	{
	'A' : 'SITEID',
	'B' : 'BSCID',
	'C' : 'CELLNAME',
	'D' : 'CELLID',
	'E' : 'MCC',
	'F' : 'MNC',
	'G' : 'LAC',
	'H' : 'NCC',
	'I' : 'BCC',
	'J' : 'BCCH',
	'K' : 'Non-main BCCH',
	'L' : 'Power',
	'M' : 'PDCH Num',
	'N' : 'SDCCH Num',
	'O' : 'HSN',
	'P' : 'Frequency List of MA Group',
	'Q' : 'Latitude Int',
	'R' : 'Latitude Decimal',
	'S' : 'Longitude Int',
	'T' : 'Longitude Decimal',
	'U' :  'Azimuth',
	'V' : 'Altitude Int',
	'W' : 'BSC OPC',
	'X' : 'BSC NSEI queue',
	'Y' : 'BVCI begin',
	# 'Z' : 'BSC NSEI2',
	# 'AA' : 'BVCI2',
	'AB' : 'TRX NUM',
	'AC' : 'TRXGROUP',
	'AD' : 'SECTOREQM NUM',
	'AE' : 'Last Character'
	})

for index in range( 1, len(gCellSiteIdList) + 1 ):
	gsmCellConfigurationWs.append(
		{
		'A' : gCellSiteIdList[index - 1],
		'B' : bscIdList[index - 1],
		'C' : gCellNameList[index - 1],
		'D' : gCellIdList[index - 1],
		'E' : gCellMccList[index - 1],
		'F' : gCellMncList[index - 1],
		'G' : gCellLacList[index - 1],
		'H' : gCellNccList[index - 1],
		'I' : gCellBccList[index - 1],
		'J' : gCellBcchList[index - 1],
		'K' : gCellFreListList[index - 1],
		'L' : gCellPowerList[index - 1],
		'M' : gCellPdchNumList[index - 1],
		'N' : gCellSdcchNumList[index - 1],
		'O' : gCellHsnList[index - 1],
		'P' : gCellFreMaGroupListList[index - 1],
		'Q' : gCellLatitudeIntList[index - 1],
		'R' : gCellLatitudeDecimalList[index - 1],
		'S' : gCellLongitudeIntList[index - 1],
		'T' : gCellLongitudeDecimalList[index - 1],
		'U' : gCellAzimuthList[index - 1],
		'V' : gCellAltitudeIntList[index - 1],
		'W' : bscOpcList[index - 1],
		# 'X' : gCellNseiList[index - 1],
		# 'Y' : gcellbvciList[index - 1],
		# 'Z' : gCellNsei2List[index - 1],
		# 'AA' : gcellbvci2List[index - 1],
		'AB' : gCellTrxNumList[index - 1],
		'AC' : gCellTrxGroupList[index - 1],
		'AD' : gCellSectorEqmNumList[index - 1],
		'AE' : gCellLastCharacterList[index - 1]
		})

umtsCellConfigurationWs.append(
	{
	'A' : 'SITEID',
	'B' : 'RNCID',
	'C' : 'CELLNAME',
	'D' : 'CELLID',
	'E' : 'LAC',
	'F' : 'SAC',
	'G' : 'RAC',
	'H' : 'URA',
	'I' : 'UARFCN',
	'J' : 'PSC',
	'K' : 'MaxDlPower',
	'L' : 'PCPICH',
	'M' : 'TCELL',
	'N' : 'Latitude',
	'O' : 'Longitude',
	'P' : 'Azimuth',
	'Q' : 'Height',
	'R' : 'Sector ID',
	'S' : 'SectorEqm ID',
	'T' : 'BasebandEqm ID',
	'U' : 'Uplink UARFCN',
	'V' : 'BandIndicator',
	'W' : 'FrequencyBandwidth',
	'X' : 'Logical RNCID',
	'Y' : 'Max Coverage',
	'Z' : 'Antenna Opening',
	'AA' : 'electricalTilt',
	'AB' : 'mechanicalTilt'
	})
print "uCellSiteIdListlen: ", len(uCellSiteIdList)
print uCellSiteIdList

for index in range( 1, len(uCellSiteIdList) + 1 ):
	umtsCellConfigurationWs.append(
		{
		'A' : uCellSiteIdList[index - 1],
		'B' : rncIdList[index - 1],
		'C' : uCellNameList[index - 1],
		'D' : uCellIdList[index - 1],
		'E' : uCellLacList[index - 1],
		'F' : uCellSacList[index - 1],
		'G' : uCellRacList[index - 1],
		'H' : uCellUra,
		'I' : uCellUarfcnList[index - 1],
		'J' : uCellPscList[index - 1],
		'K' : uCellMaxDlPowerList[index - 1],
		'L' : uCellPcpichList[index - 1],
		'M' : uCellTcellList[index - 1],
		'N' : uCellLatitudeList[index - 1],
		'O' : uCellLongitudeList[index - 1],
		'P' : uCellAzimuthList[index - 1],
		'Q' : uCellHeightList[index - 1],
		'R' : uCellSectorIdList[index - 1],
		'S' : uCellSectorEqmIdList[index - 1],
		'T' : uCellBasebandEqmIdList[index - 1],
		'U' : uCellUplinkUarfcnList[index - 1],
		'V' : uCellBandIndicatorList[index - 1],
		'W' : uCellFrequencyBandwidthList[index - 1],
		'X' : logicalRncIdList[index - 1],
		'Y' : maxCoverage,
		'Z' : antennaOpening,
		'AA' : uCellElectricalTiltList[index - 1],
		'AB' : uCellMechanicalTiltList[index - 1]
 		})

eptCellInfoWb.save('input\\ept info.xlsx')



