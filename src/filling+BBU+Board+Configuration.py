#-*- coding:utf-8 -*




from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

import os
import time

#----------------
#select siteIDList/WBBPNumList/GTMUNumList/UBRINumList from the board info Excel
#----------------
boardInfoWb = load_workbook(filename = r'input\\board info.xlsx')
print "Worksheet name(s):", boardInfoWb.get_sheet_names()

boardInfoWsIndex = 0
boardInfoWs = boardInfoWb.get_sheet_by_name(boardInfoWb.get_sheet_names()[boardInfoWsIndex])
print "Work Sheet Title: ", boardInfoWs.title

siteIDList = []
WBBPNumList = []
GTMUNumList = []
UBRINumList = []

siteIDColIndex = 'A'
WBBPNumColIndex = 'B'
GTMUNumColIndex = 'C'
UBRINumColIndex = 'D'

print boardInfoWs.get_highest_row()
for index in range(2, boardInfoWs.get_highest_row() + 1):
    siteIDList.append( boardInfoWs.cell(siteIDColIndex + str(index)).value )
    WBBPNumList.append( boardInfoWs.cell(WBBPNumColIndex + str(index)).value )
    GTMUNumList.append( boardInfoWs.cell(GTMUNumColIndex + str(index)).value )
    UBRINumList.append( boardInfoWs.cell(UBRINumColIndex + str(index)).value )

print len(siteIDList)
print len(WBBPNumList)
print len(GTMUNumList)
print len(UBRINumList)

#----------------
#write BBU Board Configuration into SummaryData Excel
#----------------
BBUBoardConfigurationWb = Workbook()
UBRIWs = BBUBoardConfigurationWb.create_sheet(0,'Baseband Radio Interface')
GTMUWs = BBUBoardConfigurationWb.create_sheet(1,'GTMU')
BBPWs = BBUBoardConfigurationWb.create_sheet(2,'BBPList')
BasebandWs = BBUBoardConfigurationWb.create_sheet(3, 'BASEBANDEQMList')

UBRIContent = ['0','0','0','UBRI','UNBLOCKED']
print UBRIContent
UBRIWs.append(
    {
    'A' : 'Base Station',
    'B' : 'Baseband Radio Interface'
    })
UBRIWs.merge_cells('B1:H1')

UBRIWs.append(
    {
    'A':'*Name',
    'B':'Cabinet No.',
    'C':'Subrack No.',
    'D':'Slot No.',
    'E':'Board Type',
    'F':'Administrative State'
    })
for index in range( 1, len(UBRINumList)+1 ):
    if( UBRINumList[index - 1] != 0 ):
        UBRIWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : UBRIContent[0],
            'C' : UBRIContent[1],
            'D' : UBRIContent[2],
            'E' : UBRIContent[3],
            'F' : UBRIContent[4]
            })

GTMUContent = ['0','0','6','GTMU','UNBLOCKED']
print GTMUContent
GTMUWs.append(
    {
    'A' : 'Base Station',
    'B' : 'Evolved GTMU as Baseband Radio Interface'
    })
GTMUWs.merge_cells('B1:H1')

GTMUWs.append(
    {
    'A':'*Name',
    'B':'Cabinet No.',
    'C':'Subrack No.',
    'D':'Slot No.',
    'E':'Board Type',
    'F':'Administrative State'
    })
for index in range( 1, len(GTMUNumList)+1 ):
    if( GTMUNumList[index - 1] != 0 ):
        GTMUWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : GTMUContent[0],
            'C' : GTMUContent[1],
            'D' : GTMUContent[2],
            'E' : GTMUContent[3],
            'F' : GTMUContent[4]
            })

BBPCommonContent = ['0','0','WBBP','FDD','UNBLOCKED','FULL']
# priority from top to end : 2-1-4-5 if has UBRI
BBPIfUBRIExistContent = ['2','1','4','5']
# priority from top to end : 2-1-4-0 if has GTMU
BBPIfGTMUExistContent = ['2','1','4','0']
# priority from top to end : 2-1-4-0 if doesn't have GTMU nor UBRI
BBPIfUmtsOnlyContent = ['2','1','4','0'] 
BBPContent = ['0','0','2','1','4','5','2','1','4','0','WBBP','FDD','UNBLOCKED','FULL']
print BBPContent
BBPWs.append(
    {
    'A' : 'Base Station',
    'B' : 'Base Band Processor'
    })
BBPWs.merge_cells('B1:N1')

BBPWs.append(
    {
    'A':'*Name',
    'B':'Cabinet No.',
    'C':'Subrack No.',
    'D':'Slot No.',
    'E':'Board Type',
    'F':'Work Mode',
    'G':'Administrative State',
    'J':'Hardware Capacity Enhance'
    })
for index in range( 1, len(siteIDList)+1 ):
    if( UBRINumList[index - 1] != 0 ):
        for i in range(WBBPNumList[index - 1]):
            BBPWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BBPCommonContent[0],
                'C' : BBPCommonContent[1],
                'D' : BBPIfUBRIExistContent[i],
                'E' : BBPCommonContent[2],
                'F' : BBPCommonContent[3],
                'G' : BBPCommonContent[4],
                'J' : BBPCommonContent[5]
                })
    elif( GTMUNumList[index - 1] != 0 ):
        for i in range(int(WBBPNumList[index - 1])):
            BBPWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BBPCommonContent[0],
                'C' : BBPCommonContent[1],
                'D' : BBPIfGTMUExistContent[i],
                'E' : BBPCommonContent[2],
                'F' : BBPCommonContent[3],
                'G' : BBPCommonContent[4],
                'J' : BBPCommonContent[5]
                })
    elif( UBRINumList[index - 1] == 0 and  GTMUNumList[index - 1] == 0 ):
        for i in range(int(WBBPNumList[index - 1])):
            BBPWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BBPCommonContent[0],
                'C' : BBPCommonContent[1],
                'D' : BBPIfUmtsOnlyContent[i],
                'E' : BBPCommonContent[2],
                'F' : BBPCommonContent[3],
                'G' : BBPCommonContent[4],
                'J' : BBPCommonContent[5]
                })        

BasebandUlContent = ['0','1','UL','DEM_2_CHAN','0,0,2','0,0,2;0,0,1','0,0,4','0,0,4;0,0,5','0,0,4;0,0,0']
BasebandDlContent = ['0','1','DL','0,0,2','0,0,2;0,0,1','0,0,4','0,0,4;0,0,5','0,0,4;0,0,0']
print BasebandUlContent
print BasebandDlContent
BasebandWs.append(
    {
    'A' : 'Base Station',
    'B' : 'Baseband Equipment'
    })
BasebandWs.merge_cells('B1:E1')

BasebandWs.append(
    {
    'A' : '*Name',
    'B' : 'Baseband Equipment ID',
    'C' : 'Baseband Equipment Type',
    'D' : 'UMTS UL Demodulation Mode',
    'E' : 'Baseband Equipment Board'
    })
for index in range( 1, len(siteIDList) + 1 ):
    num = WBBPNumList[index - 1]
    if( num == 0 ):
        continue
    elif( num == 1 ):
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandUlContent[0],
            'C' : BasebandUlContent[2],
            'D' : BasebandUlContent[3],
            'E' : BasebandUlContent[4]
            })
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandDlContent[0],
            'C' : BasebandDlContent[2],
            'E' : BasebandDlContent[3]
            })
    elif( num == 2 ):
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandUlContent[0],
            'C' : BasebandUlContent[2],
            'D' : BasebandUlContent[3],
            'E' : BasebandUlContent[5]
            })
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandDlContent[0],
            'C' : BasebandDlContent[2],
            'E' : BasebandDlContent[4]
            })        
    elif( num == 3 ):
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandUlContent[0],
            'C' : BasebandUlContent[2],
            'D' : BasebandUlContent[3],
            'E' : BasebandUlContent[5]
            })
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandDlContent[0],
            'C' : BasebandDlContent[2],
            'E' : BasebandDlContent[4]
            })    
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandUlContent[1],
            'C' : BasebandUlContent[2],
            'D' : BasebandUlContent[3],
            'E' : BasebandUlContent[6]
            })
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandDlContent[1],
            'C' : BasebandDlContent[2],
            'E' : BasebandDlContent[5]
            }) 
    elif( num == 4 ):
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandUlContent[0],
            'C' : BasebandUlContent[2],
            'D' : BasebandUlContent[3],
            'E' : BasebandUlContent[5]
            })
        BasebandWs.append(
            {
            'A' : siteIDList[index - 1],
            'B' : BasebandDlContent[0],
            'C' : BasebandDlContent[2],
            'E' : BasebandDlContent[4]
            })    
        if( UBRINumList[index - 1] != 0 ):
            BasebandWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BasebandUlContent[1],
                'C' : BasebandUlContent[2],
                'D' : BasebandUlContent[3],
                'E' : BasebandUlContent[7]
                })
            BasebandWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BasebandDlContent[1],
                'C' : BasebandDlContent[2],
                'E' : BasebandDlContent[6]
                }) 
        elif( GTMUNumList[index - 1] != 0 ):
            BasebandWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BasebandUlContent[1],
                'C' : BasebandUlContent[2],
                'D' : BasebandUlContent[3],
                'E' : BasebandUlContent[8]
                })
            BasebandWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BasebandDlContent[1],
                'C' : BasebandDlContent[2],
                'E' : BasebandDlContent[7]
                })
        elif( UBRINumList[index - 1] == 0 and GTMUNumList[index - 1] == 0 ):  
            BasebandWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BasebandUlContent[1],
                'C' : BasebandUlContent[2],
                'D' : BasebandUlContent[3],
                'E' : BasebandUlContent[8]
                })
            BasebandWs.append(
                {
                'A' : siteIDList[index - 1],
                'B' : BasebandDlContent[1],
                'C' : BasebandDlContent[2],
                'E' : BasebandDlContent[7]
                })


folder = time.strftime(r"BBU_Board_%Y-%m-%d_%H-%M-%S",time.localtime())
os.makedirs(r'%s/%s'%(os.getcwd(),folder))

BBUBoardConfigurationWb.save(folder + '\\BBU Board Configuration.xlsx')

