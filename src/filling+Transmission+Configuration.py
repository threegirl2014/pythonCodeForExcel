#-*- coding:utf-8 -*




from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

import os
import time

currentSiteWsIndex  = input("scenario 0 : L -> GUL\nscenario 1 : L -> UL\nscenario 2 : GU\nscenario 3 : UO\ninput the scenario: ")

#----------------
#select siteIds & IPs From IP Planning File
#----------------
IPPlanningWb = load_workbook(filename = r'input\\ip planning.xlsx')
print "Worksheet name(s):", IPPlanningWb.get_sheet_names()

IPPlanningWsIndex = 0
IPPlanningWs = IPPlanningWb.get_sheet_by_name(IPPlanningWb.get_sheet_names()[IPPlanningWsIndex])
print "Work Sheet Title: ", IPPlanningWs.title

siteIdList = []

egbtsDEVIpList = []
egbtsVLANIDList = []
egbtsGWIpList = []

bscIpPoolIpIndexList = []
rncIpPoolIpIndexList = []

nodebDEVIpList = []
nodebVLANIDList = []
nodebGWIpList = []

omIpList = []
omGWIpList = []
omVLANIDList = []

bscNameList = []
bscDEVIpList = []

rncNameList = []
rncDEVIpList = []
rncDEVNetworkIpList = []
rncDEVNetworkIpMaskList = []

rncSubrackList = []
rncSlotList = []
rncTrunkNoList = []
rncTrunkIpAddressIndexList = []
rncTrunkIpList = []
logicalRNCIdList = []

SCTPLNKIdList = []
SCTPLNK2IdList = []
NCPSCTPLNKIDList = []
CCPSCTPLNKIPList = []

OMDestinationIPList = []
ipClkList = []
ntpMasterList = []
ntpSlaveList = []


siteIdIndex = 'A'
nodebGWIpIndex = 'B'
nodebDEVIpIndex = 'C'
nodebVLANIDIndex = 'D'
omGWIpIndex = 'E'
omIpIndex = 'F'
omVLANIDIndex = 'G'
egbtsGWIpIndex = 'H'
egbtsDEVIpIndex = 'I'
egbtsVLANIDIndex = 'J'
bscNameIndex = 'K'
bscDevipIndex = 'L'
bscDevportIndex = 'O'
rncNameIndex = 'P'
rncDEVIpIndex = 'Q'
rncDEVNetworkIPIndex = 'R'
rncDEVNetworkIPMaskIndex = 'S'
rncSubrackIndex = 'T'
rncSlotIndex = 'U'
rncTrunkNoIndex = 'V'
rncTrunkIpAddressIndexListIndex = 'W'
rncTrunkIpIndex = 'X'
logicalRNCIdIndex = 'Y'
bscIpPoolIpIndex = 'Z'
rncIpPoolIpIndex = 'AA'
SCTPLNKIdIndex = 'AB'
SCTPLNK2IdIndex = 'AC'
NCPSCTPLNKIDIndex = 'AD'
CCPSCTPLNKIPIndex = 'AE'
OMDestinationIPIndex = 'AF'
ipClkIndex = 'AG'
ntpMasterIndex = 'AH'
ntpSlaveIndex = 'AI'


# for all
cabinetSubrackSlot = ['0','0','7']
# for DEVIP(BaseStaion)List
egbtsDEVIPMask = '255.255.255.248'
nodebDEVIPMask = '255.255.255.248'
# for IPRT(BaseStation)List
egbtsIPRTIndex = '20'
nodebIPRTIndex = '21'
egbtsIPRTMask = '255.255.255.255'
nodebIPRTMask = '255.255.255.248'
# for VLANMAP
egbtsVLANMask = '255.255.255.248'
nodebVLANMask = '255.255.255.248'
# for SCTPLNK(BaseStation)List
egbtsSCTPLNKNoList = ['100','101']
nodebSCTPLNKNoList = ['98','99']
# for IPPATH(BaseStation)List
egbtsIPPATHIDList = ['100','101','102']
# for basestation transport data
rncTrunkIpMask = '255.255.255.248'
omIpMask = '255.255.255.248'


for index in range( 2, IPPlanningWs.get_highest_row() + 1 ):
    siteIdList.append( IPPlanningWs.cell(siteIdIndex + str(index)).value )

    nodebDEVIpList.append( IPPlanningWs.cell(nodebDEVIpIndex + str(index)).value )
    nodebVLANIDList.append( IPPlanningWs.cell(nodebVLANIDIndex + str(index)).value )
    nodebGWIpList.append( IPPlanningWs.cell(nodebGWIpIndex + str(index)).value )
    
    omIpList.append( IPPlanningWs.cell(omIpIndex + str(index)).value )
    omVLANIDList.append( IPPlanningWs.cell(omVLANIDIndex + str(index)).value )
    omGWIpList.append( IPPlanningWs.cell(omGWIpIndex + str(index)).value )

    rncNameList.append( IPPlanningWs.cell(rncNameIndex + str(index)).value )
    rncDEVIpList.append( IPPlanningWs.cell(rncDEVIpIndex + str(index)).value )
    rncDEVNetworkIpList.append( IPPlanningWs.cell(rncDEVNetworkIPIndex + str(index)).value )
    rncDEVNetworkIpMaskList.append( IPPlanningWs.cell(rncDEVNetworkIPMaskIndex + str(index)).value )
    rncSubrackList.append( IPPlanningWs.cell(rncSubrackIndex + str(index)).value )
    rncSlotList.append( IPPlanningWs.cell(rncSlotIndex + str(index)).value )
    rncTrunkNoList.append( IPPlanningWs.cell(rncTrunkNoIndex + str(index)).value )
    rncTrunkIpAddressIndexList.append( IPPlanningWs.cell(rncTrunkIpAddressIndexListIndex + str(index)).value )
    rncTrunkIpList.append( IPPlanningWs.cell(rncTrunkIpIndex + str(index)).value )
    logicalRNCIdList.append( IPPlanningWs.cell(logicalRNCIdIndex + str(index)).value )
    NCPSCTPLNKIDList.append( IPPlanningWs.cell(NCPSCTPLNKIDIndex + str(index)).value)
    CCPSCTPLNKIPList.append( IPPlanningWs.cell(CCPSCTPLNKIPIndex + str(index)).value)

    ipClkList.append( IPPlanningWs.cell(ipClkIndex + str(index)).value )
    ntpMasterList.append( IPPlanningWs.cell(ntpMasterIndex + str(index)).value )
    ntpSlaveList.append( IPPlanningWs.cell(ntpSlaveIndex + str(index)).value )

    rncIpPoolIpIndexList.append( IPPlanningWs.cell(rncIpPoolIpIndex + str(index)).value )
    OMDestinationIPList.append( IPPlanningWs.cell(OMDestinationIPIndex + str(index)).value )

    if( currentSiteWsIndex == 0 or currentSiteWsIndex == 2 ):
        bscIpPoolIpIndexList.append( IPPlanningWs.cell(bscIpPoolIpIndex + str(index)).value )
        egbtsDEVIpList.append( IPPlanningWs.cell(egbtsDEVIpIndex + str(index)).value )
        egbtsVLANIDList.append( IPPlanningWs.cell(egbtsVLANIDIndex + str(index)).value )
        egbtsGWIpList.append( IPPlanningWs.cell(egbtsGWIpIndex + str(index)).value )
        bscNameList.append( IPPlanningWs.cell(bscNameIndex + str(index)).value )
        bscDEVIpList.append( IPPlanningWs.cell(bscDevipIndex + str(index)).value )
        SCTPLNKIdList.append( IPPlanningWs.cell(SCTPLNKIdIndex + str(index)).value )
        SCTPLNK2IdList.append( IPPlanningWs.cell(SCTPLNK2IdIndex + str(index)).value )

print len(siteIdList)

#----------------
#write Transmission Configuration into SummaryData Excel
#----------------

transmissionConfigurationWb = Workbook()
baseStationTrasportDataWs = transmissionConfigurationWb.create_sheet(0,'Base Station Trasport Data')
baseStationDEVIPWs = transmissionConfigurationWb.create_sheet(1,'DEVIP(BaseStation)List')
baseStationIPRTWs = transmissionConfigurationWb.create_sheet(2,'IPRT(BaseStation)List')
VLANMAPWs = transmissionConfigurationWb.create_sheet(3,'VLANMAP')
baseStationSCTPLNKWs = transmissionConfigurationWb.create_sheet(4,'SCTPLNK(BaseStation)List')
cpBearWs = transmissionConfigurationWb.create_sheet(5,'CPBEARERList')
gbtsAbisCPWs = transmissionConfigurationWb.create_sheet(6,'GBTSABISCPList')
iubCPWs = transmissionConfigurationWb.create_sheet(7,'IUBCPList')
baseStationIPPathWs = transmissionConfigurationWb.create_sheet(8,'IPPATH(BaseStation)List')
gbtsPathWs = transmissionConfigurationWb.create_sheet(9,'GBTSPATHList')
userPlaneHostWs = transmissionConfigurationWb.create_sheet(10,'USERPLANEHOSTList')
userPlanePeerWs = transmissionConfigurationWb.create_sheet(11,'User Plane Peer')
epGroupWs = transmissionConfigurationWb.create_sheet(12,'EPGROUPList')
iubWs = transmissionConfigurationWb.create_sheet(13,'IUBList')

if( currentSiteWsIndex == 0 ):
    baseStationTrasportDataContent = ['DBS3900','GBTS_Radio','1024','1025','100000','100000',
    'NodeB_Radio','IP_TRANS','DEDICATED','0','SCTP','3000','0','SCTP','3001','IP','3','100000','100000']
    print baseStationTrasportDataContent
    baseStationTrasportDataWs.append(
        {
        'A' : 'Base Station',
        'C' : 'Logical eGBTS',
        'D' : 'eGBTS',
        'F' : 'BSC IP address',
        'H' : 'eGBTS IP address',
        'L' : 'Abis Control Plane',
        'U' : 'NodeB',
        'W' : 'Logical NodeB',
        'AC' : 'RNC IP address',
        'AE' : 'Trunk IP Address',
        'AK' : 'NodeB IP address',
        'AO' : 'Iub Control Plane',
        'AV' : 'Iub User Plane',
        'BB' : 'OM Plane'
        })
    baseStationTrasportDataWs.merge_cells('A1:B1')
    baseStationTrasportDataWs.merge_cells('D1:E1')
    baseStationTrasportDataWs.merge_cells('F1:G1')
    baseStationTrasportDataWs.merge_cells('H1:K1')
    baseStationTrasportDataWs.merge_cells('L1:T1')
    baseStationTrasportDataWs.merge_cells('U1:V1')
    baseStationTrasportDataWs.merge_cells('W1:AB1')
    baseStationTrasportDataWs.merge_cells('AC1:AD1')
    baseStationTrasportDataWs.merge_cells('AE1:AJ1')
    baseStationTrasportDataWs.merge_cells('AK1:AN1')
    baseStationTrasportDataWs.merge_cells('AO1:AU1')
    baseStationTrasportDataWs.merge_cells('AV1:BA1')
    baseStationTrasportDataWs.merge_cells('BB1:BC1')
    
    baseStationTrasportDataWs.append(
        {
        'A' : '*Name',
        'B' : '*Product Type',
        'C' : '*BSC Name',
        'D' : 'BTS Name',
        'E' : '*GBTS Radio Template',
        'F' : '*Control Plane IP address',
        'G' : '*User  Plane IP address',
        'H' : '*Control Plane IP address',
        'I' : '*Control Plane Subnet mask',
        'J' : '*User Plane IP address',
        'K' : '*User Plane Subnet mask',
        'L' : '*SCTP Client Port',
        'M' : '*SCTP2 Client Port',
        'N' : '*SCTP link ID',
        'O' : 'SCTP2 link ID',
        'P' : 'Adjacent Node ID',
        'Q' : '*Adjacent Node Name',
        'R' : '*Transmission Resource Pool Index',
        'S' : '*Forward Bandwidth',
        'T' : '*Backward Bandwidth',
        'U' : '*NodeB Name',
        'V' : '*NodeB Radio Template',
        'W' : '*RNC Name',
        'X' : 'Logical RNC ID',
        'Y' : 'NodeB ID',
        'Z' : 'IUB Transport Bearer Type',
        'AA' : 'Sharing Type Of NodeB',
        'AB' : 'Cn Operator Index',
        'AC' : '*Control Plane IP address',
        'AD' : '*User Plane IP address',
        'AE' : 'Subrack No.',
        'AF' : 'Slot No.',
        'AG' : 'Trunk No.',
        'AH' : 'IP address index',
        'AI' : 'Local IP address',
        'AJ' : 'Subnet mask',
        'AK' : '*Control Plane IP address',
        'AL' : '*Control Plane Subnet mask',
        'AM' : '*User Plane IP address',
        'AN' : '*User Plane Subnet mask',
        'AO' : '*NCP Bearing Link Type',
        'AP' : '*NCP SCTP link ID',
        'AQ' : '*NCP SCTP Port',
        'AR' : '*CCP Port No.',
        'AS' : '*CCP Bearing Link Type',
        'AT' : '*CCP SCTP link ID',
        'AU' : '*CCP SCTP Port',
        'AV' : 'Adjacent Node ID',
        'AW' : '*Adjacent Node Name',
        'AX' : '*Transport Type',
        'AY' : '*Transmission Resource Pool Index',
        'AZ' : '*Forward Bandwidth',
        'BA' : '*Backward Bandwidth',
        'BB' : '*OM IP',
        'BC' : '*OM IP Mask'
        })
    for index in range(1, len(siteIdList)+1 ):
        baseStationTrasportDataWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : baseStationTrasportDataContent[0],
            'C' : bscNameList[index - 1],
            'D' : str(siteIdList[index - 1]) + 'GSM',
            'E' : baseStationTrasportDataContent[1],
            'F' : bscDEVIpList[index - 1],
            'G' : bscDEVIpList[index - 1],
            'H' : egbtsDEVIpList[index -1],
            'I' : egbtsDEVIPMask,
            'J' : egbtsDEVIpList[index -1],
            'K' : egbtsDEVIPMask,
            'L' : baseStationTrasportDataContent[2],
            'M' : baseStationTrasportDataContent[3],
            'N' : SCTPLNKIdList[index - 1],
            'O' : SCTPLNK2IdList[index -1],
            'P' : str(siteIdList[index - 1])[2:],
            'Q' : str(siteIdList[index - 1]) + 'GSM',
            'R' : bscIpPoolIpIndexList[index - 1],
            'S' : baseStationTrasportDataContent[4],
            'T' : baseStationTrasportDataContent[5],
            'U' : str(siteIdList[index - 1]) + 'UMTS',
            'V' : baseStationTrasportDataContent[6],
            'W' : rncNameList[index - 1],
            'X' : logicalRNCIdList[index - 1],
            'Y' : str(siteIdList[index - 1])[2:],
            'Z' : baseStationTrasportDataContent[7],
            'AA' : baseStationTrasportDataContent[8],
            'AB' : baseStationTrasportDataContent[9],
            'AC' : rncDEVIpList[index - 1],
            'AD' : rncDEVIpList[index - 1],
            'AE' : rncSubrackList[index -1],
            'AF' : rncSlotList[index - 1],
            'AG' : rncTrunkNoList[index - 1],
            'AH' : rncTrunkIpAddressIndexList[index -1],
            'AI' : rncTrunkIpList[index - 1],
            'AJ' : rncTrunkIpMask,
            'AK' : nodebDEVIpList[index - 1],
            'AL' : nodebDEVIPMask,
            'AM' : nodebDEVIpList[index - 1],
            'AN' : nodebDEVIPMask,
            'AO' : baseStationTrasportDataContent[10],
            'AP' : NCPSCTPLNKIDList[index - 1],
            'AQ' : baseStationTrasportDataContent[11],
            'AR' : baseStationTrasportDataContent[12],
            'AS' : baseStationTrasportDataContent[13],
            'AT' : CCPSCTPLNKIPList[index - 1],
            'AU' : baseStationTrasportDataContent[14],
            'AV' : str(siteIdList[index - 1])[2:],
            'AW' : str(siteIdList[index - 1]) + 'UMTS',
            'AX' : baseStationTrasportDataContent[15],
            'AY' : rncIpPoolIpIndexList[index - 1],
            'AZ' : baseStationTrasportDataContent[17],
            'BA' : baseStationTrasportDataContent[18],
            'BB' : omIpList[index - 1],
            'BC' : omIpMask
            })
elif( currentSiteWsIndex == 1 ):
    baseStationTrasportDataContent = ['DBS3900','NodeB_Radio','IP_TRANS','DEDICATED','0',
    '3000','0','3001','SCTP','SCTP','IP','3','100000','100000']
    print baseStationTrasportDataContent
    baseStationTrasportDataWs.append(
        {
        'A' : 'Base Station',
        'C' : 'NodeB',
        'E' : 'Logical NodeB',
        'K' : 'Trunk IP Address',
        'Q' : 'RNC IP address',
        'S' : 'NodeB IP address',
        'W' : 'Iub Control Plane',
        'AJ' : 'OM Plane'        
        })
    baseStationTrasportDataWs.merge_cells('A1:B1')
    baseStationTrasportDataWs.merge_cells('C1:D1')
    baseStationTrasportDataWs.merge_cells('E1:J1')
    baseStationTrasportDataWs.merge_cells('K1:P1')
    baseStationTrasportDataWs.merge_cells('Q1:R1')
    baseStationTrasportDataWs.merge_cells('S1:V1')
    baseStationTrasportDataWs.merge_cells('W1:AI1')
    baseStationTrasportDataWs.merge_cells('AJ1:AK1')
    baseStationTrasportDataWs.append(
        {
        'A' : '*Name',
        'B' : '*Type',
        'C' : '*NodeB Name',
        'D' : '*NodeB Radio Template',
        'E' : '*RNC Name',
        'F' : '*Logical RNC ID',
        'G' : 'NodeB ID',
        'H' : '*IUB Transport Bearer Type',
        'I' : '*Sharing Type Of NodeB',
        'J' : 'Cn Operator Index',
        'K' : 'Subrack No.',
        'L' : 'Slot No.',
        'M' : 'Trunk No.',
        'N' : 'IP address index',
        'O' : 'Local IP address',
        'P' : 'Subnet mask',
        'Q' : '*Control Plane IP address',
        'R' : '*User Plane IP address',
        'S' : '*Control Plane IP address',
        'T' : '*Control Plane Subnet mask',
        'U' : '*User Plane IP address',
        'V' : '*User Plane Subnet mask',
        'W' : '*NCP SCTP Port',
        'X' : '*CCP Port No.',
        'Y' : '*CCP SCTP Port',
        'Z' : '*NCP Bearing Link Type',
        'AA' : '*NCP SCTP link ID',
        'AB' : '*CCP Bearing Link Type',
        'AC' : '*CCP SCTP link ID',
        'AD' : 'Adjacent Node ID',
        'AE' : '*Adjacent Node Name',
        'AF' : '*Transport Type',
        'AG' : '*Transmission Resource Pool Index',
        'AH' : '*Forward Bandwidth',
        'AI' : '*Backward Bandwidth',
        'AJ' : '*OM IP',
        'AK' : '*OM IP Mask'       
        })
    for index in range(1, len(siteIdList)+1 ):
        baseStationTrasportDataWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : baseStationTrasportDataContent[0],
            'C' : str(siteIdList[index - 1]) + 'UMTS',
            'D' : baseStationTrasportDataContent[1],
            'E' : rncNameList[index - 1],
            'F' : logicalRNCIdList[index - 1],
            'G' : str(siteIdList[index - 1])[2:],
            'H' : baseStationTrasportDataContent[2],
            'I' : baseStationTrasportDataContent[3],
            'J' : baseStationTrasportDataContent[4],
            'K' : rncSubrackList[index -1],
            'L' : rncSlotList[index - 1],
            'M' : rncTrunkNoList[index - 1],
            'N' : rncTrunkIpAddressIndexList[index -1],
            'O' : rncTrunkIpList[index - 1],
            'P' : rncTrunkIpMask,
            'Q' : rncDEVIpList[index - 1],
            'R' : rncDEVIpList[index - 1],
            'S' : nodebDEVIpList[index - 1],
            'T' : nodebDEVIPMask,
            'U' : nodebDEVIpList[index - 1],
            'V' : nodebDEVIPMask,
            'W' : baseStationTrasportDataContent[5],
            'X' : baseStationTrasportDataContent[6],
            'Y' : baseStationTrasportDataContent[7],
            'Z' : baseStationTrasportDataContent[8],
            'AA' : NCPSCTPLNKIDList[index - 1],
            'AB' : baseStationTrasportDataContent[9],
            'AC' : CCPSCTPLNKIPList[index - 1],
            'AD' : str(siteIdList[index - 1])[2:],
            'AE' : str(siteIdList[index - 1]) + 'UMTS',
            'AF' : baseStationTrasportDataContent[10],
            'AG' : rncIpPoolIpIndexList[index - 1],
            'AH' : baseStationTrasportDataContent[12],
            'AI' : baseStationTrasportDataContent[13],
            'AJ' : omIpList[index - 1],
            'AK' : omIpMask
            })
elif( currentSiteWsIndex == 2 ):
    baseStationTrasportDataContent = ['DBS3900','GBTS_Radio','1024','1025','100000','100000','NodeB_Radio','IP_TRANS','DEDICATED',
    '0','SCTP','3000','0','SCTP','3001','IP','100000','100000','IUB','255.255.255.255','255.255.255.192','DISABLE']
    print baseStationTrasportDataContent
    baseStationTrasportDataWs.append(
        {
        'A' : '*Name',
        'B' : '*Product Type',
        'C' : '*Node Template',
        'D' : 'Base Station Configuration',
        'E' : 'Site Location',
        'F' : 'Geo-coordinate data format',
        'G' : 'LatitudeDegFormat',
        'H' : 'LatitudeSecFormat',
        'I' : 'LongitudeDegFormat',
        'J' : 'LongitudeSecFormat',
        'K' : '*BSC Name',
        'L' : '*BTS Name',
        'M' : '*GSM Radio Template',
        'N' : '*Control Plane IP address',
        'O' : '*User Plane IP address',
        'P' : '*Control Plane IP address',
        'Q' : '*Control Plane Subnet mask',
        'R' : '*User Plane IP address',
        'S' : '*User Plane Subnet mask',
        'T' : 'IP address',
        'U' : 'Subnet mask',
        'V' : 'Adjacent Node ID',
        'W' : '*Adjacent Node Name',
        'X' : '*Transmission Resource Pool Index',
        'Y' : '*SCTP Client Port',
        'Z' : '*SCTP2 Client Port',
        'AA' : '*SCTP link ID',
        'AB' : '*SCTP2 link ID',
        'AC' : '*Forward Bandwidth',
        'AD' : '*Backward Bandwidth',
        'AE' : '*NodeB Name',
        'AF' : '*UMTS Radio Template',
        'AG' : '*RNC Name',
        'AH' : 'Logical RNC ID',
        'AI' : 'NodeB ID',
        'AJ' : 'IUB Transport Bearer Type',
        'AK' : 'Sharing Type Of NodeB',
        'AL' : 'Cn Operator Index',
        'AM' : 'Validation indication',
        'AN' : 'NodeB Host Type',
        'AO' : 'Iub Type',
        'AP' : '*Control Plane IP address',
        'AQ' : '*User Plane IP address',
        'AR' : 'Subrack No.',
        'AS' : 'Slot No.',
        'AT' : 'Trunk No.',
        'AU' : 'IP address index',
        'AV' : 'Local IP address',
        'AW' : 'Subnet mask',
        'AX' : '*Control Plane IP address',
        'AY' : '*Control Plane Subnet mask',
        'AZ' : '*User Plane Host ID',
        'BA' : '*User Plane IP address',
        'BB' : '*User Plane Subnet mask',
        'BC' : 'IP address',
        'BD' : 'Subnet mask',
        'BE' : '*NCP Bearing Link Type',
        'BF' : '*NCP SCTP Port',
        'BG' : '*CCP Port No.',
        'BH' : '*CCP Bearing Link Type',
        'BI' : '*CCP SCTP Port',
        'BJ' : '*NCP SCTP link ID',
        'BK' : '*CCP SCTP link ID',
        'BL' : 'ANI',
        'BM' : 'NAME',
        'BN' : 'TRANST',
        'BO' : 'IPPOOLINDEX',
        'BP' : 'TXBW',
        'BQ' : 'RXBW',
        'BR' : 'NODET',
        'BS' : '*OM IP',
        'BT' : '*OM IP Mask',
        'BU' : 'IPClock Client IP1',
        'BV' : 'IPClock Client Mask1',
        'BW' : 'IPClock Client IP2',
        'BX' : 'IPClock Client Mask2',
        'BY' : 'Destination IP1',
        'BZ' : 'Destination Mask1',
        'CA' : 'Next Hop IP Address1',
        'CB' : 'VLAN ID1',
        'CC' : 'Destination IP2',
        'CD' : 'Destination Mask2',
        'CE' : 'Next Hop IP Address2',
        'CF' : 'VLAN ID2',
        'CG' : 'Destination IP3',
        'CH' : 'Destination Mask3',
        'CI' : 'Next Hop IP Address3',
        'CJ' : 'VLAN ID3',
        'CK' : 'Destination IP4',
        'CL' : 'Destination Mask4',
        'CM' : 'Next Hop IP Address4',
        'CN' : 'VLAN ID4',
        'CO' : 'Destination IP5',
        'CP' : 'Destination Mask5',
        'CQ' : 'Next Hop IP Address5',
        'CR' : 'VLAN ID5',
        'CS' : 'Destination IP6',
        'CT' : 'Destination Mask6',
        'CU' : 'Next Hop IP Address6',
        'CV' : 'VLAN ID6',
        'CW' : 'Destination IP7',
        'CX' : 'Destination Mask7',
        'CY' : 'Next Hop IP Address7',
        'CZ' : 'VLAN ID7',
        'DA' : 'Destination IP8',
        'DB' : 'Destination Mask8',
        'DC' : 'Next Hop IP Address8',
        'DD' : 'VLAN ID8',
        'DE' : 'Destination IP9',
        'DF' : 'Destination Mask9',
        'DG' : 'Next Hop IP Address9',
        'DH' : 'VLAN ID9',
        'DI' : 'Destination IP10',
        'DJ' : 'Destination Mask10',
        'DK' : 'Next Hop IP Address10',
        'DL' : 'VLAN ID10',
        'DM' : 'Destination IP11',
        'DN' : 'Destination Mask11',
        'DO' : 'Next Hop IP Address11',
        'DP' : 'VLAN ID11',
        'DQ' : 'SCTP Fault Real-time Reporting Switch'            
        })
    for index in range(1, len(siteIdList) + 1):
        baseStationTrasportDataWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : baseStationTrasportDataContent[0],
            'K' : bscNameList[index - 1],
            'L' : str(siteIdList[index - 1]) + 'GSM',
            'M' : baseStationTrasportDataContent[1],   
            'N' : bscDEVIpList[index - 1],
            'O' : bscDEVIpList[index - 1],
            'P' : egbtsDEVIpList[index - 1],
            'Q' : egbtsDEVIPMask,
            'R' : egbtsDEVIpList[index - 1],
            'S' : egbtsDEVIPMask,
            'T' : egbtsDEVIpList[index - 1],
            'U' : egbtsDEVIPMask,
            'V' : str(siteIdList[index - 1])[2:],
            'W' : str(siteIdList[index - 1]) + 'GSM',
            'X' : bscIpPoolIpIndexList[index - 1],
            'Y' : baseStationTrasportDataContent[2],
            'Z' : baseStationTrasportDataContent[3],      
            'AA' : SCTPLNKIdList[index - 1],
            'AB' : SCTPLNK2IdList[index - 1],
            'AC' : baseStationTrasportDataContent[4],
            'AD' : baseStationTrasportDataContent[5],
            'AE' : str(siteIdList[index - 1]) + 'UMTS',
            'AF' : baseStationTrasportDataContent[6],
            'AG' : rncNameList[index - 1],
            'AH' : logicalRNCIdList[index - 1],
            'AI' : str(siteIdList[index - 1])[2:],
            'AJ' : baseStationTrasportDataContent[7],
            'AK' : baseStationTrasportDataContent[8],
            'AL' : baseStationTrasportDataContent[9],
            'AP' : rncDEVIpList[index - 1],
            'AQ' : rncDEVIpList[index - 1],
            'AR' : rncSubrackList[index - 1],
            'AS' : rncSlotList[index - 1],
            'AT' : rncTrunkNoList[index - 1],
            'AU' : rncTrunkIpAddressIndexList[index -1],
            'AV' : rncTrunkIpList[index - 1],
            'AW' : rncTrunkIpMask,
            'AX' : nodebDEVIpList[index - 1],
            'AY' : nodebDEVIPMask,
            'AZ' : logicalRNCIdList[index - 1],
            'BA' : nodebDEVIpList[index - 1],
            'BB' : nodebDEVIPMask,
            'BC' : nodebDEVIpList[index - 1],
            'BD' : nodebDEVIPMask,
            'BE' : baseStationTrasportDataContent[10],
            'BF' : baseStationTrasportDataContent[11],
            'BG' : baseStationTrasportDataContent[12],
            'BH' : baseStationTrasportDataContent[13],
            'BI' : baseStationTrasportDataContent[14],
            'BJ' : NCPSCTPLNKIDList[index - 1],
            'BK' : CCPSCTPLNKIPList[index - 1],
            'BL' : str(siteIdList[index - 1])[2:],
            'BM' : str(siteIdList[index - 1]) + 'UMTS',
            'BN' : baseStationTrasportDataContent[15],
            'BO' : rncIpPoolIpIndexList[index - 1],
            'BP' : baseStationTrasportDataContent[16],
            'BQ' : baseStationTrasportDataContent[17],
            'BR' : baseStationTrasportDataContent[18],
            'BS' : omIpList[index - 1],
            'BT' : omIpMask,
            'BU' : omIpList[index - 1],
            'BV' : omIpMask,
            'BW' : omIpList[index - 1],
            'BX' : omIpMask,
            'BY' : bscDEVIpList[index - 1],
            'BZ' : baseStationTrasportDataContent[19],
            'CA' : egbtsGWIpList[index - 1],
            'CB' : egbtsVLANIDList[index - 1],
            'CC' : rncDEVNetworkIpList[index - 1],
            'CD' : rncDEVNetworkIpMaskList[index - 1],
            'CE' : nodebGWIpList[index - 1],
            'CF' : nodebVLANIDList[index - 1],
            'CG' : OMDestinationIPList[index - 1],
            'CH' : baseStationTrasportDataContent[20],
            'CI' : omGWIpList[index - 1],
            'CJ' : omVLANIDList[index - 1],
            'DA' : ipClkList[index - 1],
            'DB' : baseStationTrasportDataContent[19],
            'DC' : omGWIpList[index - 1],
            'DD' : omVLANIDList[index - 1],
            'DI' : ntpMasterList[index - 1],
            'DJ' : baseStationTrasportDataContent[19],
            'DK' : omGWIpList[index - 1],
            'DL' : omVLANIDList[index - 1],
            'DM' : ntpSlaveList[index -1],
            'DN' : baseStationTrasportDataContent[19],
            'DO' : omGWIpList[index - 1],
            'DP' : omVLANIDList[index - 1],
            'DQ' : baseStationTrasportDataContent[21]
            })

elif( currentSiteWsIndex == 3 ):
    baseStationTrasportDataContent = ['DBS3900','NodeB_Radio','IP_TRANS','DEDICATED',
    '0','SCTP','3000','0','SCTP','3001','IP','100000','100000','255.255.255.248','10.191.155.0','255.255.255.192','255.255.255.255','DISABLE']
    print baseStationTrasportDataContent
    baseStationTrasportDataWs.append(
        {
        'A' : '*Name',
        'B' : '*Product Type',
        'C' : '*Node Template',
        'D' : 'Base Station Configuration',
        'E' : '*NodeB Name',
        'F' : '*UMTS Radio Template',
        'G' : '*RNC Name',
        'H' : 'Logical RNC ID',
        'I' : 'NodeB ID',
        'J' : 'IUB Transport Bearer Type',
        'K' : 'Sharing Type Of NodeB',
        'L' : 'Cn Operator Index',
        'M' : '*Control Plane IP address',
        'N' : '*User Plane IP address',
        'O' : 'Subrack No.',
        'P' : 'Slot No.',
        'Q' : 'Trunk No.',
        'R' : 'IP address index',
        'S' : 'Local IP address',
        'T' : 'Subnet mask',
        'U' : '*Control Plane IP address',
        'V' : '*Control Plane Subnet mask',
        'W' : '*User Plane Host ID',
        'X' : '*User Plane IP address',
        'Y' : '*User Plane Subnet mask',
        'Z' : 'IP address',
        'AA' : 'Subnet mask',
        'AB' : '*NCP Bearing Link Type',
        'AC' : '*NCP SCTP Port',
        'AD' : '*CCP Port No.',
        'AE' : '*CCP Bearing Link Type',
        'AF' : '*CCP SCTP Port',
        'AG' : '*NCP SCTP link ID',
        'AH' : '*CCP SCTP link ID',
        'AI' : 'Adjacent Node ID',
        'AJ' : '*Adjacent Node Name',
        'AK' : '*Transport Type',
        'AL' : '*Transmission Resource Pool Index',
        'AM' : '*Forward Bandwidth',
        'AN' : '*Backward Bandwidth',
        'AO' : '*OM IP',
        'AP' : '*OM IP Mask',
        'AQ' : 'IPClock Client IP1',
        'AR' : 'IPClock Client Mask1',
        'AS' : 'IPClock Client IP2',
        'AT' : 'IPClock Client Mask2',
        'AU' : 'Cable Length(m)',
        'AV' : 'Destination IP1',
        'AW' : 'Destination Mask1',
        'AX' : 'Next Hop IP Address1',
        'AY' : 'VLAN ID1',
        'AZ' : 'Destination IP2',
        'BA' : 'Destination Mask2',
        'BB' : 'Next Hop IP Address2',
        'BC' : 'VLAN ID2',
        'BD' : 'Destination IP3',
        'BE' : 'Destination Mask3',
        'BF' : 'Next Hop IP Address3',
        'BG' : 'VLAN ID3',
        'BH' : 'Destination IP4',
        'BI' : 'Destination Mask4',
        'BJ' : 'Next Hop IP Address4',
        'BK' : 'VLAN ID4',
        'BL' : 'Destination IP5',
        'BM' : 'Destination Mask5',
        'BN' : 'Next Hop IP Address5',
        'BO' : 'VLAN ID5',
        'BP' : 'Destination IP6',
        'BQ' : 'Destination Mask6',
        'BR' : 'Next Hop IP Address6',
        'BS' : 'VLAN ID6',
        'BT' : 'Destination IP7',
        'BU' : 'Destination Mask7',
        'BV' : 'Next Hop IP Address7',
        'BW' : 'VLAN ID7',
        'BX' : 'Destination IP8',
        'BY' : 'Destination Mask8',
        'BZ' : 'Next Hop IP Address8',
        'CA' : 'VLAN ID8',
        'CB' : 'Destination IP9',
        'CC' : 'Destination Mask9',
        'CD' : 'Next Hop IP Address9',
        'CE' : 'VLAN ID9',
        'CF' : 'Destination IP10',
        'CG' : 'Destination Mask10',
        'CH' : 'Next Hop IP Address10',
        'CI' : 'VLAN ID10',
        'CJ' : 'Destination IP11',
        'CK' : 'Destination Mask11',
        'CL' : 'Next Hop IP Address11',
        'CM' : 'VLAN ID11',
        'CN' : 'SCTP Fault Real-time Reporting Switch',
        'CO' : 'Site Location',
        'CP' : 'Geo-coordinate data format',
        'CQ' : 'LatitudeDegFormat',
        'CR' : 'LatitudeSecFormat',
        'CS' : 'LongitudeDegFormat',
        'CT' : 'LongitudeSecFormat'       
        })
    for index in range(1, len(siteIdList) + 1):
        baseStationTrasportDataWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : baseStationTrasportDataContent[0],
            'E' : str(siteIdList[index - 1]) + 'UMTS',
            'F' : baseStationTrasportDataContent[1],
            'G' : rncNameList[index - 1],
            'H' : logicalRNCIdList[index - 1],
            'I' : str(siteIdList[index - 1])[2:],
            'J' : baseStationTrasportDataContent[2],
            'K' : baseStationTrasportDataContent[3],
            'L' : baseStationTrasportDataContent[4],
            'M' : rncDEVIpList[index - 1],
            'N' : rncDEVIpList[index - 1],
            'O' : rncSubrackList[index -1],
            'P' : rncSlotList[index - 1],
            'Q' : rncTrunkNoList[index - 1],
            'R' : rncTrunkIpAddressIndexList[index -1],
            'S' : rncTrunkIpList[index - 1],
            'T' : rncTrunkIpMask,
            'U' : nodebDEVIpList[index - 1],
            'V' : nodebDEVIPMask,
            'W' : logicalRNCIdList[index - 1],
            'X' : nodebDEVIpList[index - 1],
            'Y' : nodebDEVIPMask,
            'Z' : nodebDEVIpList[index - 1],
            'AA' : nodebDEVIPMask,
            'AB' : baseStationTrasportDataContent[5],
            'AC' : baseStationTrasportDataContent[6],
            'AD' : baseStationTrasportDataContent[7],
            'AE' : baseStationTrasportDataContent[8],
            'AF' : baseStationTrasportDataContent[9],
            'AG' : NCPSCTPLNKIDList[index - 1],
            'AH' : CCPSCTPLNKIPList[index - 1],
            'AI' : str(siteIdList[index - 1])[2:],
            'AJ' : str(siteIdList[index - 1]) + 'UMTS',
            'AK' : baseStationTrasportDataContent[10],
            'AL' : rncIpPoolIpIndexList[index - 1],
            'AM' : baseStationTrasportDataContent[11],
            'AN' : baseStationTrasportDataContent[12],
            'AO' : omIpList[index - 1],
            'AP' : omIpMask,
            'AQ' : omIpList[index - 1],
            'AR' : omIpMask,
            'AS' : omIpList[index - 1],
            'AT' : omIpMask,
            'AZ' : rncDEVNetworkIpList[index - 1],
            'BA' : rncDEVNetworkIpMaskList[index - 1],
            'BB' : nodebGWIpList[index - 1],
            'BC' : nodebVLANIDList[index - 1],
            'BD' : OMDestinationIPList[index - 1],
            'BE' : baseStationTrasportDataContent[15],
            'BF' : omGWIpList[index - 1],
            'BG' : omVLANIDList[index - 1],
            'BX' : ipClkList[index - 1],
            'BY' : baseStationTrasportDataContent[16],
            'BZ' : omGWIpList[index - 1],
            'CA' : omVLANIDList[index - 1],           
            # 'CC' : baseStationTrasportDataContent[16],
            # 'CD' : omGWIpList[index - 1],
            # 'CE' : omVLANIDList[index - 1],
            'CF' : ntpMasterList[index - 1],
            'CG' : baseStationTrasportDataContent[16],
            'CH' : omGWIpList[index - 1],
            'CI' : omVLANIDList[index - 1],
            'CJ' : ntpSlaveList[index - 1],
            'CK' : baseStationTrasportDataContent[16],
            'CL' : omGWIpList[index - 1],
            'CM' : omVLANIDList[index - 1],
            'CN' : baseStationTrasportDataContent[17]
            })    



baseStationDEVIPContent = ['BASE_BOARD','ETH','0','0','GSM','UMTS']
print baseStationDEVIPContent
baseStationDEVIPWs.append(
    {
    'A' : 'Base Station',
    'B' : 'Device IP Address'
    })
baseStationDEVIPWs.merge_cells('B1:K1')

baseStationDEVIPWs.append(
    {
    'A' : '*Name',
    'B' : 'Cabinet No.',
    'C' : 'Subrack No.',
    'D' : 'Slot No.',
    'E' : 'Subboard Type',
    'F' : 'Port Type',
    'G' : 'Port No.',
    'H' : 'VRF Index',
    'I' : 'IP Address',
    'J' : 'Mask',
    'K' : 'User Label'
    })
for index in range( 1, len(siteIdList)+1 ):
    if( currentSiteWsIndex == 0 ):
        baseStationDEVIPWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : cabinetSubrackSlot[0],
            'C' : cabinetSubrackSlot[1],
            'D' : cabinetSubrackSlot[2],
            'E' : baseStationDEVIPContent[0],
            'F' : baseStationDEVIPContent[1],
            'G' : baseStationDEVIPContent[2],
            'H' : baseStationDEVIPContent[3],
            'I' : egbtsDEVIpList[index - 1],
            'J' : egbtsDEVIPMask,
            'K' : baseStationDEVIPContent[4]
            })
    baseStationDEVIPWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : cabinetSubrackSlot[0],
        'C' : cabinetSubrackSlot[1],
        'D' : cabinetSubrackSlot[2],
        'E' : baseStationDEVIPContent[0],
        'F' : baseStationDEVIPContent[1],
        'G' : baseStationDEVIPContent[2],
        'H' : baseStationDEVIPContent[3],
        'I' : nodebDEVIpList[index - 1],
        'J' : nodebDEVIPMask,
        'K' : baseStationDEVIPContent[5]
        })
    


baseStationIPRTContent = ['20','21','BASE_BOARD','NEXTHOP','0','GSM','UMTS']
print baseStationIPRTContent
baseStationIPRTWs.append(
    {
    'A' : 'Base Station',
    'B' : 'IP Route'
    })
baseStationIPRTWs.merge_cells('B1:O1')

baseStationIPRTWs.append(
    {
    'A' : '*Name',
    'B' : 'Route Index',
    'C' : 'Cabinet No.',
    'D' : 'Subrack No.',
    'E' : 'Slot No.',
    'F' : 'Subboard Type',
    'G' : 'Route Type',
    'H' : 'Port Type',
    'I' : 'Port No.',
    'J' : 'VRF Index',
    'K' : 'Destination IP',
    'L' : 'Mask',
    'M' : 'Next Hop IP',
    'N' : 'Preference',
    'O' : 'Description Info'
    })
for index in range( 1, len(siteIdList)+1 ):
    if( currentSiteWsIndex == 0 ):
        baseStationIPRTWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : baseStationIPRTContent[0],
            'C' : cabinetSubrackSlot[0],
            'D' : cabinetSubrackSlot[1],
            'E' : cabinetSubrackSlot[2],
            'F' : baseStationIPRTContent[2],
            'G' : baseStationIPRTContent[3],
            'J' : baseStationIPRTContent[4],
            'K' : bscDEVIpList[index - 1],
            'L' : egbtsIPRTMask,
            'M' : egbtsGWIpList[index - 1],
            'O' : baseStationIPRTContent[5]
            })
    baseStationIPRTWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : baseStationIPRTContent[1],
        'C' : cabinetSubrackSlot[0],
        'D' : cabinetSubrackSlot[1],
        'E' : cabinetSubrackSlot[2],
        'F' : baseStationIPRTContent[2],
        'G' : baseStationIPRTContent[3],
        'J' : baseStationIPRTContent[4],
        'K' : rncDEVNetworkIpList[index - 1],
        'L' : nodebIPRTMask,
        'M' : nodebGWIpList[index - 1],
        'O' : baseStationIPRTContent[6]
        })


VLANMAPContent = ['0','SINGLEVLAN','DISABLE']
print VLANMAPContent
VLANMAPWs.append(
    {
    'A' : 'Base Station',
    'B' : 'VLAN Mapping Based on Next Hop IP'
    })
VLANMAPWs.merge_cells('B1:I1')

VLANMAPWs.append(
    {
    'A' : '*Name',
    'B' : 'VRF Index',
    'C' : 'Next Hop IP',
    'D' : 'Mask',
    'E' : 'VLAN Mode',
    'F' : 'VLAN ID',
    'G' : 'Set VLAN Priority'
    })
for index in range( 1, len(siteIdList)+1 ):
    if( currentSiteWsIndex == 0 ):
        VLANMAPWs.append(
            {
            'A' : siteIdList[index -1],
            'B' : VLANMAPContent[0],
            'C' : egbtsGWIpList[index - 1],
            'D' : egbtsVLANMask,
            'E' : VLANMAPContent[1],
            'F' : egbtsVLANIDList[index - 1],
            'G' : VLANMAPContent[2]
            })
    VLANMAPWs.append(
        {
        'A' : siteIdList[index -1],
        'B' : VLANMAPContent[0],
        'C' : nodebGWIpList[index - 1],
        'D' : nodebVLANMask,
        'E' : VLANMAPContent[1],
        'F' : nodebVLANIDList[index - 1],
        'G' : VLANMAPContent[2]
        })
  

baseStationSCTPLNKContent = ['1024','1025','3000','3001','58080','G-CP1','G-CP2','U-CP1','U-CP2']
print baseStationSCTPLNKContent
baseStationSCTPLNKWs.append(
    {
    'A' : 'Base Station',
    'B' : 'SCTP Link'
    })
baseStationSCTPLNKWs.merge_cells('B1:AB1')
baseStationSCTPLNKWs.append(
    {
    'A' : '*Name',
    'B' : 'Link No.',
    'C' : 'Cabinet No.',
    'D' : 'Subrack No.',
    'E' : 'Slot No.',
    'F' : 'Maximum Stream No.',
    'G' : 'Control Mode',
    'H' : 'VRF Index',
    'I' : 'First Local IP Address',
    'J' : 'Second Local IP Address',
    'K' : 'Local SCTP Port No.',
    'L' : 'First Peer IP Address',
    'M' : 'Second Peer IP Address',
    'N' : 'Peer SCTP Port No.',
    'O' : 'RTO Min Value(ms)',
    'P' : 'RTO Max Value(ms)',
    'Q' : 'RTO Initial Value(ms)',
    'R' : 'RTO Alpha Value',
    'S' : 'RTO Beta Value',
    'T' : 'Heart-beat Interval(ms)',
    'U' : 'Max Association Retransmission',
    'V' : 'Max Path Retransmission',
    'W' : 'Checksum Arithmetic Type',
    'X' : 'Switch Back Flag',
    'Y' : 'Heart-beat Times When Switch Back',
    'Z' : 'Block Flag',
    'AA' : 'SACK Timeout(ms)',
    'AB' : 'Description Info'
    })
for index in range( 1, len(siteIdList)+1 ):
    if( currentSiteWsIndex == 0 ):
        baseStationSCTPLNKWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsSCTPLNKNoList[0],
            'C' : cabinetSubrackSlot[0],
            'D' : cabinetSubrackSlot[1],
            'E' : cabinetSubrackSlot[2],
            'I' : egbtsDEVIpList[index - 1],
            'K' : baseStationSCTPLNKContent[0],
            'L' : bscDEVIpList[index - 1],
            'N' : baseStationSCTPLNKContent[4],
            'AB' : baseStationSCTPLNKContent[5]
            })
        baseStationSCTPLNKWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsSCTPLNKNoList[1],
            'C' : cabinetSubrackSlot[0],
            'D' : cabinetSubrackSlot[1],
            'E' : cabinetSubrackSlot[2],
            'I' : egbtsDEVIpList[index - 1],
            'K' : baseStationSCTPLNKContent[1],
            'L' : bscDEVIpList[index - 1],
            'N' : baseStationSCTPLNKContent[4],
            'AB' : baseStationSCTPLNKContent[6]
            })
    baseStationSCTPLNKWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : nodebSCTPLNKNoList[0],
        'C' : cabinetSubrackSlot[0],
        'D' : cabinetSubrackSlot[1],
        'E' : cabinetSubrackSlot[2],
        'I' : nodebDEVIpList[index - 1],
        'K' : baseStationSCTPLNKContent[2],
        'L' : rncDEVIpList[index - 1],
        'N' : baseStationSCTPLNKContent[4],
        'AB' : baseStationSCTPLNKContent[7]
        })
    baseStationSCTPLNKWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : nodebSCTPLNKNoList[1],
        'C' : cabinetSubrackSlot[0],
        'D' : cabinetSubrackSlot[1],
        'E' : cabinetSubrackSlot[2],
        'I' : nodebDEVIpList[index - 1],
        'K' : baseStationSCTPLNKContent[3],
        'L' : rncDEVIpList[index - 1],
        'N' : baseStationSCTPLNKContent[4],
        'AB' : baseStationSCTPLNKContent[8]
        })

cpBearContent = ['MASTER','SCTP','AUTO_MODE']
print cpBearContent
cpBearWs.append(
    {
    'A' : 'Base Station',
    'B' : 'Control Port Bearer'
    })
cpBearWs.merge_cells('B1:F1')

cpBearWs.append(
    {
    'A' : '*Name',
    'B' : 'CP Bear No.',
    'C' : 'Flag',
    'D' : 'Bear Type',
    'E' : 'Link No.',
    'F' : 'Control Mode'
    })
for index in range(1, len(siteIdList)+1 ):
    if( currentSiteWsIndex == 0 ):
        cpBearWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsSCTPLNKNoList[0],
            'C' : cpBearContent[0],
            'D' : cpBearContent[1],
            'E' : egbtsSCTPLNKNoList[0],
            'F' : cpBearContent[2]
            })
        cpBearWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsSCTPLNKNoList[1],
            'C' : cpBearContent[0],
            'D' : cpBearContent[1],
            'E' : egbtsSCTPLNKNoList[1],
            'F' : cpBearContent[2]
            })
    cpBearWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : nodebSCTPLNKNoList[0],
        'C' : cpBearContent[0],
        'D' : cpBearContent[1],
        'E' : nodebSCTPLNKNoList[0],
        'F' : cpBearContent[2]
        })
    cpBearWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : nodebSCTPLNKNoList[1],
        'C' : cpBearContent[0],
        'D' : cpBearContent[1],
        'E' : nodebSCTPLNKNoList[1],
        'F' : cpBearContent[2]
        })

if( currentSiteWsIndex == 0 ):
    gbtsAbisCPContent = ['0']
    print gbtsAbisCPContent
    gbtsAbisCPWs.append(
        {
        'A' : 'eGBTS',
        'B' : 'GBTS Abis Control Port'
        })
    gbtsAbisCPWs.merge_cells('B1:C1')
    
    gbtsAbisCPWs.append(
        {
        'A' : '*BTS Name',
        'B' : 'Control Port ID',
        'C' : 'CP Bear No.'
        })
    for index in range(1, len(siteIdList)+1):
        gbtsAbisCPWs.append(
            {
            'A' : str(siteIdList[index - 1]) + 'GSM',
            'B' : gbtsAbisCPContent[0],
            'C' : egbtsSCTPLNKNoList[0]
            })
        gbtsAbisCPWs.append(
            {
            'A' : str(siteIdList[index - 1]) + 'GSM',
            'B' : gbtsAbisCPContent[0],
            'C' : egbtsSCTPLNKNoList[1]
            })
    

iubCpContent = ['NCP','CCP','0','MASTER']
print iubCpContent
iubCPWs.append(
    {
    'A' : 'NodeB',
    'B' : 'NodeB Iub Control Port'
    })
iubCPWs.merge_cells('B1:E1')

iubCPWs.append(
    {
    'A' : '*NodeB Name',
    'B' : 'Port Type',
    'C' : 'CP Port No.',
    'D' : 'Belong Flag',
    'E' : 'CPBear ID'
    })
for index in range(1, len(siteIdList)+1):
    iubCPWs.append(
        {
        'A' : str(siteIdList[index - 1]) + 'UMTS',
        'B' : iubCpContent[0],
        'D' : iubCpContent[3],
        'E' : nodebSCTPLNKNoList[0]
        })
    iubCPWs.append(
        {
        'A' : str(siteIdList[index - 1]) + 'UMTS',
        'B' : iubCpContent[1],
        'C' : iubCpContent[2],
        'D' : iubCpContent[3],
        'E' : nodebSCTPLNKNoList[1]
        })
    
if( currentSiteWsIndex == 0 ):
    baseStationIPPathContent = ['BASE_BOARD','ETH','0','DISABLE','FIXED',
    '26','34','46','0','HQ','GSM-UP1','GSM-UP2','GSM-UP3']
    print baseStationIPPathContent
    baseStationIPPathWs.append(
        {
        'A' : 'Base Station',
        'B' : 'IP Path'
        })
    baseStationIPPathWs.merge_cells('B1:U1')
    
    baseStationIPPathWs.append(
        {
        'A' : '*Name',
        'B' : 'Path ID',
        'C' : 'Cabinet No.',
        'D' : 'Subrack No.',
        'E' : 'Slot No.',
        'F' : 'Subboard Type',
        'G' : 'Port Type',
        'H' : 'Port No.',
        'I' : 'Join Transmission Resource Group',
        'J' : 'Transmission Resource Group ID',
        'K' : 'Path Type',
        'L' : 'DSCP',
        'M' : 'VRF Index',
        'N' : 'Local IP',
        'O' : 'Peer IP',
        'P' : 'Transport Resource Type',
        'Q' : 'IPMUX Switch Flag',
        'R' : 'Max Subframe Length(byte)',
        'S' : 'Max Frame Length(byte)',
        'T' : 'Max Timer(ms)',
        'U' : 'Description'
        })
    for index in range(1, len(siteIdList)+1 ):
        baseStationIPPathWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsIPPATHIDList[0],
            'C' : cabinetSubrackSlot[0],
            'D' : cabinetSubrackSlot[1],
            'E' : cabinetSubrackSlot[2],
            'F' : baseStationIPPathContent[0],
            'G' : baseStationIPPathContent[1],
            'H' : baseStationIPPathContent[2],
            'I' : baseStationIPPathContent[3],
            'K' : baseStationIPPathContent[4],
            'L' : baseStationIPPathContent[5],
            'M' : baseStationIPPathContent[8],
            'N' : egbtsDEVIpList[index - 1],
            'O' : bscDEVIpList[index - 1],
            'P' : baseStationIPPathContent[9],
            'U' : baseStationIPPathContent[10],
            })
        baseStationIPPathWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsIPPATHIDList[1],
            'C' : cabinetSubrackSlot[0],
            'D' : cabinetSubrackSlot[1],
            'E' : cabinetSubrackSlot[2],
            'F' : baseStationIPPathContent[0],
            'G' : baseStationIPPathContent[1],
            'H' : baseStationIPPathContent[2],
            'I' : baseStationIPPathContent[3],
            'K' : baseStationIPPathContent[4],
            'L' : baseStationIPPathContent[6],
            'M' : baseStationIPPathContent[8],
            'N' : egbtsDEVIpList[index - 1],
            'O' : bscDEVIpList[index - 1],
            'P' : baseStationIPPathContent[9],
            'U' : baseStationIPPathContent[11],
            })
        baseStationIPPathWs.append(
            {
            'A' : siteIdList[index - 1],
            'B' : egbtsIPPATHIDList[2],
            'C' : cabinetSubrackSlot[0],
            'D' : cabinetSubrackSlot[1],
            'E' : cabinetSubrackSlot[2],
            'F' : baseStationIPPathContent[0],
            'G' : baseStationIPPathContent[1],
            'H' : baseStationIPPathContent[2],
            'I' : baseStationIPPathContent[3],
            'K' : baseStationIPPathContent[4],
            'L' : baseStationIPPathContent[7],
            'M' : baseStationIPPathContent[8],
            'N' : egbtsDEVIpList[index - 1],
            'O' : bscDEVIpList[index - 1],
            'P' : baseStationIPPathContent[9],
            'U' : baseStationIPPathContent[12],
            })
    
if( currentSiteWsIndex == 0 ):
    gbtsPathWs.append(
        {
        'A' : 'eGBTS',
        'B' : 'GBTS IP Path'
        })
    gbtsPathWs.append(
        {
        'A' : '*BTS Name',
        'B' : 'Path ID'
        })
    for index in range(1, len(siteIdList)+1 ):
        gbtsPathWs.append(
            {
            'A' : str(siteIdList[index - 1]) + 'GSM',
            'B' : egbtsIPPATHIDList[0]
            })
        gbtsPathWs.append(
            {
            'A' : str(siteIdList[index - 1]) + 'GSM',
            'B' : egbtsIPPATHIDList[1]
            })
        gbtsPathWs.append(
            {
            'A' : str(siteIdList[index - 1]) + 'GSM',
            'B' : egbtsIPPATHIDList[2]
            })

    
    
userPlaneHostContent = ['0','IPv4','DISABLE','UMTS_User_plan']
print userPlaneHostContent
userPlaneHostWs.append(
    {
    'A' : 'Base Station',
    'B' : 'User Plane Host'
    })
userPlaneHostWs.merge_cells('B1:I1')

userPlaneHostWs.append(
    {
    'A' : '*Name',
    'B' : 'User Plane Host ID',
    'C' : 'VRF Index',
    'D' : 'IP Version',
    'E' : 'Local IP Address',
    'F' : 'Local IPv6 Address',
    'G' : 'IPSec Switch',
    'H' : 'Security Host ID',
    'I' : 'User Label'
    })
for index in range(1, len(siteIdList)+1 ):
    userPlaneHostWs.append(
        {
        'A' : siteIdList[index - 1],
        'B' : logicalRNCIdList[index - 1],
        'C' : userPlaneHostContent[0],
        'D' : userPlaneHostContent[1],
        'E' : nodebDEVIpList[index - 1],
        'G' : userPlaneHostContent[2],
        'I' : userPlaneHostContent[3]
        })
    

userPlanePeerContent = ['30','0','IPv4','DISABLE','AUTO_MODE']
print userPlanePeerContent
userPlanePeerWs.append(
    {
    'A' : 'Base Station',
    'B' : 'User Plane Peer'
    })
userPlanePeerWs.merge_cells('B1:K1')

userPlanePeerWs.append(
    {
    'A' : '*Name',
    'B' : 'User Plane Peer ID',
    'C' : 'VRF Index',
    'D' : 'IP Version',
    'E' : 'Peer IP Address',
    'F' : 'Peer IPv6 Address',
    'G' : 'IPSec Switch',
    'H' : 'Security Peer ID',
    'I' : 'Remote ID',
    'J' : 'Control Mode',
    'K' : 'User Label'
    })
for index in range(1, len(siteIdList)+1 ):
    userPlanePeerWs.append(
        {
        'A' : siteIdList[index -1],
        'B' : userPlanePeerContent[0],
        'C' : userPlanePeerContent[1],
        'D' : userPlanePeerContent[2],
        'E' : rncDEVIpList[index - 1],
        'G' : userPlanePeerContent[3],
        'J' : userPlanePeerContent[4]
        })


epGroupContent = ['0','30']
print epGroupContent
epGroupWs.append(
    {
    'A' : 'Base Station',
    'B' : 'End Point Group'
    })
epGroupWs.merge_cells('B1:I1')

epGroupWs.append(
    {
    'A' : '*Name',
    'B' : 'End Point Group ID',
    'C' : 'VRF Index',
    'D' : 'SCTP Host ID List',
    'E' : 'SCTP Peer ID List',
    'F' : 'User Plane Host ID List',
    'G' : 'User Plane Peer ID List',
    'H' : 'Packet Filter Switch',
    'I' : 'User Label'
    })
for index in range(1, len(siteIdList)+1 ):
    epGroupWs.append(
        {
        'A' : siteIdList[index -1],
        'B' : logicalRNCIdList[index - 1],
        'C' : epGroupContent[0],
        'F' : logicalRNCIdList[index - 1],
        'G' : epGroupContent[1]
        })

iubContent = ['0']
print iubContent
iubWs.append(
    {
    'A' : 'NodeB',
    'B' : 'Iub Object'
    })
iubWs.merge_cells('B1:D1')
iubWs.append(
    {
    'A' : '*NodeB Name',
    'B' : 'Iub ID',
    'C' : 'EndPoint Group ID',
    'D' : 'User Label'
    })
for index in range(1, len(siteIdList)+1 ):
    iubWs.append(
        {
        'A' : str(siteIdList[index - 1]) + 'UMTS',
        'B' : iubContent[0],
        'C' : logicalRNCIdList[index - 1]
        })


folder = time.strftime(r"transmission_%Y-%m-%d_%H-%M-%S",time.localtime())
os.makedirs(r'%s/%s'%(os.getcwd(),folder))

if currentSiteWsIndex == 0:
    folder = folder + '\\GUL_'
elif currentSiteWsIndex == 1:
    folder = folder + '\\UL_'
elif currentSiteWsIndex == 2:
    folder = folder + '\\GU_'
elif currentSiteWsIndex == 3:
    folder = folder + '\\UO_'
transmissionConfigurationWb.save( folder + 'transmission Configuration.xlsx' )

