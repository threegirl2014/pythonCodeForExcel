#-*- coding:utf-8 -*





from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

import os
import time

#----------------
#[0][] is for UO850
#[1][] is for GU1900
#[2][] is for GO1900
#[3][] is for UO1900
#----------------
rruChainContent = [
[6,'CHAIN','COLD','LOCALPORT','0','0','2',0,'255','255','AUTO'],
[3,'LOADBALANCE','0','0',0,0,'0','0','2',3,'255','255','AUTO'],
[3,'CHAIN','COLD','LOCALPORT','0','0',0,0,'255','255','AUTO'],
[12,'CHAIN','COLD','LOCALPORT','0','0','2',3,'255','255','AUTO']
]

rruContent = [
['0',80,'0','TRUNK',6,'0','MRRU','UO850','25','15','UO','2','1','0','0','OFF'],
['0',70,'0','TRUNK',3,'0','MRRU','GU1900','25','15','GU','4','2','0','0','OFF'],
['0',70,'0','TRUNK',3,'0','MRRU','GO1900','25','15','GO','4','2','0','0','OFF'],
['0',90,'0','TRUNK',12,'0','MRRU','UO1900','25','15','UO','4','2','0','0','OFF']
]

sectorContent = [
[4,'UO850','UO850','65535'],
[7,'GU1900','GU1900','65535'],
[7,'GO1900','GO1900','65535'],
[10,'UO1900','UO1900','65535']
]

sectorEqmContent = [
[4,4],
[[7,7],[10,7],[13,7]],
[7,7],
[16,10]
]

def addUO850(siteId, rruChainWs, rruWs, sectorWs, sectorEqmWs, index):
	rruChainWs.append(
		{
		'A' : siteId,
		'B' : rruChainContent[0][0] + index,
		'C' : rruChainContent[0][1],
		'D' : rruChainContent[0][2],
		'E' : rruChainContent[0][3],
		'F' : rruChainContent[0][4],
		'G' : rruChainContent[0][5],
		'H' : rruChainContent[0][6],
		'I' : rruChainContent[0][7] + index,
		'N' : rruChainContent[0][8],
		'O' : rruChainContent[0][9],
		'P' : rruChainContent[0][10]
		})
	rruWs.append(
		{
		'A' : siteId,
		'B' : rruContent[0][0],
		'C' : rruContent[0][1] + index,
		'D' : rruContent[0][2],
		'E' : rruContent[0][3],
		'F' : rruContent[0][4] + index,
		'G' : rruContent[0][5],
		'H' : rruContent[0][6],
		'I' : rruContent[0][7],
		'J' : rruContent[0][8],
		'K' : rruContent[0][9],
		'L' : rruContent[0][10],
		'M' : rruContent[0][11],
		'N' : rruContent[0][12],
		'O' : rruContent[0][13],
		'P' : rruContent[0][14],
		'AA' : rruContent[0][15]
		})
	sectorConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorContent[0][0] + index,
		'C' : sectorContent[0][1],
		'D' : sectorContent[0][2],
		'F' : sectorContent[0][3],
		'G' : "0," + str(index + 80) + ",0,R0A;0," + str(index + 80) + ",0,R0B"				
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[0][0] + index,
		'C' : sectorEqmContent[0][1] + index,
		'D' : "0," + str(index + 80) +",0,R0A,RXTX_MODE,MASTER;0," + str(index + 80) + ",0,R0B,RX_MODE,"
		})

def addGU1900_GO1900(siteId, rruChainWs, rruWs, sectorWs, sectorEqmWs, gtmuOrUbri, index):
	if( gtmuOrUbri == 1 ):
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[1][0] + index,
			'C' : rruChainContent[1][1],
			'F' : rruChainContent[1][2],
			'G' : rruChainContent[1][3],
			'H' : rruChainContent[1][4] + 6,
			'I' : rruChainContent[1][5] + index,
			'J' : rruChainContent[1][6],
			'K' : rruChainContent[1][7],
			'L' : rruChainContent[1][8],
			'M' : rruChainContent[1][9] + index,
			'N' : rruChainContent[1][10],
			'O' : rruChainContent[1][11],
			'P' : rruChainContent[1][12]					
			})
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[2][0] + 6 + index,
			'C' : rruChainContent[2][1],
			'D' : rruChainContent[2][2],
			'E' : rruChainContent[2][3],
			'F' : rruChainContent[2][4],
			'G' : rruChainContent[2][5],
			'H' : rruChainContent[2][6] + 6,
			'I' : rruChainContent[2][7] + 3 + index,
			'N' : rruChainContent[2][8],
			'O' : rruChainContent[2][9],
			'P' : rruChainContent[2][10]					
			})
	else:
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[1][0] + index,
			'C' : rruChainContent[1][1],
			'F' : rruChainContent[1][2],
			'G' : rruChainContent[1][3],
			'H' : rruChainContent[1][4],
			'I' : rruChainContent[1][5] + index,
			'J' : rruChainContent[1][6],
			'K' : rruChainContent[1][7],
			'L' : rruChainContent[1][8],
			'M' : rruChainContent[1][9] + index,
			'N' : rruChainContent[1][10],
			'O' : rruChainContent[1][11],
			'P' : rruChainContent[1][12]
			})
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[2][0] + 6 + index,
			'C' : rruChainContent[2][1],
			'D' : rruChainContent[2][2],
			'E' : rruChainContent[2][3],
			'F' : rruChainContent[2][4],
			'G' : rruChainContent[2][5],
			'H' : rruChainContent[2][6],
			'I' : rruChainContent[2][7] + 3 + index,
			'N' : rruChainContent[2][8],
			'O' : rruChainContent[2][9],
			'P' : rruChainContent[2][10]					
			})
	rruConfigurationWb.append(
		{
		'A' : siteId,
		'B' : rruContent[1][0],
		'C' : rruContent[1][1] + index,
		'D' : rruContent[1][2],
		'E' : rruContent[1][3],
		'F' : rruContent[1][4] + index,
		'G' : rruContent[1][5],
		'H' : rruContent[1][6],
		'I' : rruContent[1][7],
		'J' : rruContent[1][8],
		'K' : rruContent[1][9],
		'L' : rruContent[1][10],
		'M' : rruContent[1][11],
		'N' : rruContent[1][12],
		'O' : rruContent[1][13],
		'P' : rruContent[1][14],
		'AA' : rruContent[1][15]
		})
	rruConfigurationWb.append(
		{
		'A' : siteId,
		'B' : rruContent[2][0],
		'C' : rruContent[2][1] + 3 + index,
		'D' : rruContent[2][2],
		'E' : rruContent[2][3],
		'F' : rruContent[2][4] + 6 + index,
		'G' : rruContent[2][5],
		'H' : rruContent[2][6],
		'I' : rruContent[2][7],
		'J' : rruContent[2][8],
		'K' : rruContent[2][9],
		'L' : rruContent[2][10],
		'M' : rruContent[2][11],
		'N' : rruContent[2][12],
		'O' : rruContent[2][13],
		'P' : rruContent[2][14],
		'AA' : rruContent[2][15]
		})
	sectorConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorContent[1][0] + index,
		'C' : sectorContent[1][1],
		'D' : sectorContent[1][2],
		'F' : sectorContent[1][3],
		'G' : "0," + str(index + 70) + ",0,R0A;0," + str(index + 70) + ",0,R0B;0," 
		+ str(index + 70) + ",0,R0C;0," + str(index + 70) + ",0,R0D;0," + str(index + 73) + ",0,R0A;0," + str(index + 73) + ",0,R0B",
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[1][0][0] + index,
		'C' : sectorEqmContent[1][0][1] + index,
		'D' : "0," + str(index + 70) +",0,R0A,RXTX_MODE,MASTER;0," + str(index + 70) + ",0,R0C,RX_MODE,"
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[1][1][0] + index,
		'C' : sectorEqmContent[1][1][1] + index,
		'D' : "0," + str(index + 70) +",0,R0B,RXTX_MODE,MASTER;0," + str(index + 70) + ",0,R0D,RX_MODE,"
		})	
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[1][2][0] + index,
		'C' : sectorEqmContent[1][2][1] + index,
		'D' : "0," + str(index + 73) +",0,R0A,RXTX_MODE,MASTER;0," + str(index + 73) + ",0,R0B,RX_MODE,"
		})

def addGU1900(siteId, rruChainWs, rruWs, sectorWs, sectorEqmWs, gtmuOrUbri, index):
	if( gtmuOrUbri == 1 ):
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[1][0] + index,
			'C' : rruChainContent[1][1],
			'F' : rruChainContent[1][2],
			'G' : rruChainContent[1][3],
			'H' : rruChainContent[1][4] + 6,
			'I' : rruChainContent[1][5] + index,
			'J' : rruChainContent[1][6],
			'K' : rruChainContent[1][7],
			'L' : rruChainContent[1][8],
			'M' : rruChainContent[1][9] + index,
			'N' : rruChainContent[1][10],
			'O' : rruChainContent[1][11],
			'P' : rruChainContent[1][12]					
			})
	else:
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[1][0] + index,
			'C' : rruChainContent[1][1],
			'F' : rruChainContent[1][2],
			'G' : rruChainContent[1][3],
			'H' : rruChainContent[1][4],
			'I' : rruChainContent[1][5] + index,
			'J' : rruChainContent[1][6],
			'K' : rruChainContent[1][7],
			'L' : rruChainContent[1][8],
			'M' : rruChainContent[1][9] + index,
			'N' : rruChainContent[1][10],
			'O' : rruChainContent[1][11],
			'P' : rruChainContent[1][12]
			})
	rruConfigurationWb.append(
		{
		'A' : siteId,
		'B' : rruContent[1][0],
		'C' : rruContent[1][1] + index,
		'D' : rruContent[1][2],
		'E' : rruContent[1][3],
		'F' : rruContent[1][4] + index,
		'G' : rruContent[1][5],
		'H' : rruContent[1][6],
		'I' : rruContent[1][7],
		'J' : rruContent[1][8],
		'K' : rruContent[1][9],
		'L' : rruContent[1][10],
		'M' : rruContent[1][11],
		'N' : rruContent[1][12],
		'O' : rruContent[1][13],
		'P' : rruContent[1][14],
		'AA' : rruContent[1][15]
		})
	sectorConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorContent[1][0] + index,
		'C' : sectorContent[1][1],
		'D' : sectorContent[1][2],
		'F' : sectorContent[1][3],
		'G' : "0," + str(index + 70) + ",0,R0A;0," + str(index + 70) + ",0,R0B;0," 
		+ str(index + 70) + ",0,R0C;0," + str(index + 70) + ",0,R0D"
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[1][0][0] + index,
		'C' : sectorEqmContent[1][0][1] + index,
		'D' : "0," + str(index + 70) +",0,R0A,RXTX_MODE,MASTER;0," + str(index + 70) + ",0,R0C,RX_MODE,"
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[1][1][0] + index,
		'C' : sectorEqmContent[1][1][1] + index,
		'D' : "0," + str(index + 70) +",0,R0B,RXTX_MODE,MASTER;0," + str(index + 70) + ",0,R0D,RX_MODE,"
		})			

def addGO1900(siteId, rruChainWs, rruWs, sectorWs, sectorEqmWs, gtmuOrUbri, index):
	if( gtmuOrUbri == 1 ):
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[2][0] + index,
			'C' : rruChainContent[2][1],
			'D' : rruChainContent[2][2],
			'E' : rruChainContent[2][3],
			'F' : rruChainContent[2][4],
			'G' : rruChainContent[2][5],
			'H' : rruChainContent[2][6] + 6,
			'I' : rruChainContent[2][7] + index,
			'N' : rruChainContent[2][8],
			'O' : rruChainContent[2][9],
			'P' : rruChainContent[2][10]					
			})
	else:
		rruChainConfigurationWb.append(
			{
			'A' : siteId,
			'B' : rruChainContent[2][0] + index,
			'C' : rruChainContent[2][1],
			'D' : rruChainContent[2][2],
			'E' : rruChainContent[2][3],
			'F' : rruChainContent[2][4],
			'G' : rruChainContent[2][5],
			'H' : rruChainContent[2][6],
			'I' : rruChainContent[2][7] + index,
			'N' : rruChainContent[2][8],
			'O' : rruChainContent[2][9],
			'P' : rruChainContent[2][10]					
			})
	rruConfigurationWb.append(
		{
		'A' : siteId,
		'B' : rruContent[2][0],
		'C' : rruContent[2][1] + index,
		'D' : rruContent[2][2],
		'E' : rruContent[2][3],
		'F' : rruContent[2][4] + index,
		'G' : rruContent[2][5],
		'H' : rruContent[2][6],
		'I' : rruContent[2][7],
		'J' : rruContent[2][8],
		'K' : rruContent[2][9],
		'L' : rruContent[2][10],
		'M' : rruContent[2][11],
		'N' : rruContent[2][12],
		'O' : rruContent[2][13],
		'P' : rruContent[2][14],
		'AA' : rruContent[2][15]
		})
	sectorConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorContent[2][0] + index,
		'C' : sectorContent[2][1],
		'D' : sectorContent[2][2],
		'F' : sectorContent[2][3],
		'G' : "0," + str(index + 70) + ",0,R0A;0," + str(index + 70) + ",0,R0B;0," 
		+ str(index + 70) + ",0,R0C;0," + str(index + 70) + ",0,R0D"
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[2][0] + index,
		'C' : sectorEqmContent[2][1] + index,
		'D' : "0," + str(index + 70) +",0,R0A,RXTX_MODE,MASTER;0," + str(index + 70) + ",0,R0B,RX_MODE,"
		})

def addUO1900(siteId, rruChainWs, rruWs, sectorWs, sectorEqmWs, index):
	rruChainWs.append(
		{
		'A' : siteId,
		'B' : rruChainContent[3][0] + index,
		'C' : rruChainContent[3][1],
		'D' : rruChainContent[3][2],
		'E' : rruChainContent[3][3],
		'F' : rruChainContent[3][4],
		'G' : rruChainContent[3][5],
		'H' : rruChainContent[3][6],
		'I' : rruChainContent[3][7] + index,
		'N' : rruChainContent[3][8],
		'O' : rruChainContent[3][9],
		'P' : rruChainContent[3][10]
		})
	rruWs.append(
		{
		'A' : siteId,
		'B' : rruContent[3][0],
		'C' : rruContent[3][1] + index,
		'D' : rruContent[3][2],
		'E' : rruContent[3][3],
		'F' : rruContent[3][4] + index,
		'G' : rruContent[3][5],
		'H' : rruContent[3][6],
		'I' : rruContent[3][7],
		'J' : rruContent[3][8],
		'K' : rruContent[3][9],
		'L' : rruContent[3][10],
		'M' : rruContent[3][11],
		'N' : rruContent[3][12],
		'O' : rruContent[3][13],
		'P' : rruContent[3][14],
		'AA' : rruContent[3][15]
		})
	sectorConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorContent[3][0] + index,
		'C' : sectorContent[3][1],
		'D' : sectorContent[3][2],
		'F' : sectorContent[3][3],
		'G' : "0," + str(index + 90) + ",0,R0A;0," + str(index + 90) + ",0,R0B;0," 
		+ str(index + 90) + ",0,R0C;0," + str(index + 90) + ",0,R0D"		
		})
	sectorEqmConfigurationWb.append(
		{
		'A' : siteId,
		'B' : sectorEqmContent[3][0] + index,
		'C' : sectorEqmContent[3][1] + index,
		'D' : "0," + str(index + 90) +",0,R0A,RXTX_MODE,MASTER;0," + str(index + 90) + ",0,R0B,RX_MODE,"
		})




rruWb = load_workbook( filename = r'input\\rru info.xlsx')
print "Worksheet name(s): ", rruWb.get_sheet_names()

rruWs = rruWb['rru info']
print "Worksheet Title: ", rruWs.title

boardInfoWb = load_workbook( filename = r'input\\board info.xlsx')
print "Worksheet name(s): ", boardInfoWb.get_sheet_names()

boardInfoWs = boardInfoWb['board info']
print "Work Sheet Title: ", boardInfoWs.title

siteIdList = []
rruScenarioSectorList = [[],[],[]]
gtmuOrUbri = []

for index in range( 2, rruWs.get_highest_row() + 1 ):
	siteIdList.append( rruWs.cell( row = index, column = 1).value )
	rruScenarioSectorList[0].append( rruWs.cell( row = index, column = 2 ).value )
	rruScenarioSectorList[1].append( rruWs.cell( row = index, column = 3 ).value )
	rruScenarioSectorList[2].append( rruWs.cell( row = index, column = 4 ).value )
	gtmuOrUbri.append( boardInfoWs.cell( row = index, column = 3 ).value )


rruAndSectorConfigurationWs = Workbook()
rruChainConfigurationWb = rruAndSectorConfigurationWs.create_sheet(0,'RRUCHAINList')
rruConfigurationWb = rruAndSectorConfigurationWs.create_sheet(1,'RRUList')
sectorConfigurationWb = rruAndSectorConfigurationWs.create_sheet(2,'SECTORList')
sectorEqmConfigurationWb = rruAndSectorConfigurationWs.create_sheet(3,'SECTOREQMList')

rruChainConfigurationWb.append(
	{
	'A' : 'Base Station',
	'B' : 'RRU Chain'
	})
rruChainConfigurationWb.merge_cells('B1:Q1')
rruChainConfigurationWb.append(
	{
	'A' : '*Name',
	'B' : 'Chain No.',
	'C' : 'Topo Type',
	'D' : 'Backup Mode',
	'E' : 'Access Type',
	'F' : 'Head Cabinet No.',
	'G' : 'Head Subrack No.',
	'H' : 'Head Slot No.',
	'I' : 'Head Port No.',
	'J' : 'Tail Cabinet No.',
	'K' : 'Tail Subrack No.',
	'L' : 'Tail Slot No.',
	'M' : 'Tail Port No.',
	'N' : 'BreakPoint Position1',
	'O' : 'BreakPoint Position2',
	'P' : 'CPRI Line Rate(Gbit/s)',
	'Q' : 'Local Slot No.'
	})
rruConfigurationWb.append(
	{
	'A' : 'Base Station',
	'B' : 'Remote Radio Unit'
	})
rruConfigurationWb.merge_cells('B1:AA1')
rruConfigurationWb.append(
	{
	'A' : '*Name',
	'B' : 'Cabinet No.',
	'C' : 'Subrack No.',
	'D' : 'Slot No.',
	'E' : 'RRU Topo Position',
	'F' : 'RRU Chain No.',
	'G' : 'RRU Position',
	'H' : 'RRU type',
	'I' : 'RRU Name',
	'J' : 'VSWR alarm post-processing threshold(0.1)',
	'K' : 'VSWR alarm threshold(0.1)',
	'L' : 'RF Unit Working Mode',
	'M' : 'Number of RX channels',
	'N' : 'Number of TX channels',
	'O' : 'IfOffset(100KHz)',
	'P' : 'RF Desensitivity(dB)',
	'Q' : 'Frequency Min Bandwidth(kHz)',
	'R' : 'Low Current Protect Switch',
	'S' : 'ALD Reuse Flag',
	'T' : 'RU Specification',
	'U' : 'RF Connect Type',
	'V' : 'Cabinet No. of RRU2',
	'W' : 'Subrack No. of RRU2',
	'X' : 'Slot No. of RRU2',
	'Y' : 'PA Efficiency Improvement Switch',
	'Z' : 'Administrative State',
	'AA' : 'VSWR alarm post-processing switch'	
	})
sectorConfigurationWb.append(
	{
	'A' : 'Base Station',
	'B' : 'Sector'	
	})
sectorConfigurationWb.merge_cells('B1:G1')
sectorConfigurationWb.append(
	{
	'A' : '*Name',
	'B' : 'Sector ID',
	'C' : 'Sector Name',
	'D' : 'Location Name',
	'E' : 'User Label',
	'F' : 'Antenna Azimuth(0.1degree)',
	'G' : 'Sector Antenna'	
	})
sectorEqmConfigurationWb.append(
	{
	'A' : 'Base Station',
	'B' : 'Sector Equipment'	
	})
sectorEqmConfigurationWb.merge_cells('B1:D1')
sectorEqmConfigurationWb.append(
	{
	'A' : '*Name',
	'B' : 'Sector Equipment ID',
	'C' : 'Sector ID',
	'D' : 'Sector Equipment Antenna'	
	})



for index in range( 1, len(siteIdList) + 1 ):
	for i in range(3):
		if( rruScenarioSectorList[ i ][ index - 1 ] == 13 ):
			addUO850(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)
			addGU1900_GO1900(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, gtmuOrUbri[ index - 1 ], i)
		elif( rruScenarioSectorList[ i ][ index - 1 ] == 14 ):
			addUO850(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)
			addGU1900(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, gtmuOrUbri[ index - 1 ], i)
		elif( rruScenarioSectorList[ i ][ index - 1 ] == 15 ):
			addUO850(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)
			addUO1900(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)	
		elif( rruScenarioSectorList[ i ][ index - 1 ] == 16 ):
			addGO1900(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, gtmuOrUbri[ index - 1 ], i)
			addUO850(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)
		elif( rruScenarioSectorList[ i ][ index - 1 ] == 24 ):
			addUO850(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)
		elif( rruScenarioSectorList[ i ][ index - 1 ] == 29 ):
			addUO850(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)
			addGO1900(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, gtmuOrUbri[ index - 1 ], i)
			addUO1900(siteIdList[index - 1], rruChainConfigurationWb, rruConfigurationWb, sectorConfigurationWb, sectorEqmConfigurationWb, i)		

folder = time.strftime(r"RRU_SECTOR_%Y-%m-%d_%H-%M-%S",time.localtime())
os.makedirs(r'%s/%s'%(os.getcwd(),folder))

rruAndSectorConfigurationWs.save( folder + '\\RRU and Sector Configuration.xlsx' )