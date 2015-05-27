#-*- coding:utf-8 -*





from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

import os
import time

def findSiteId(rawSiteId):
	length = len(rawSiteId)
	if length == 4:
		siteId = rawSiteId
	elif length > 4:
		siteId = rawSiteId[-4:]
	elif length < 4:
		siteId = '0'*(4-length) + rawSiteId
	return siteId

currentSiteWsIndex  = input("scenario 0 : L -> GUL\nscenario 1 : L -> UL\nscenario 2 : GU\nscenario 3 : UO\nscenario 4 : add UCELL in RNP(for offline-CME)\nscenario 5 : add UCELL in RNP(for online-CME)\ninput the scenario: ")

#----------------
#select cell information From EPT File
#----------------
eptWb = load_workbook( filename = r'input\\ept info.xlsx')
print "Worksheet name(s): ", eptWb.get_sheet_names()

umtsCellWs = eptWb['UMTS Cell']
gsmCellWs = eptWb['GSM Cell']
print "Worksheet Title: ", umtsCellWs.title
print "Worksheet Title: ", gsmCellWs.title

rruWb = load_workbook( filename = r'input\\rru info.xlsx')
print "Worksheet name(s): ", rruWb.get_sheet_names()

rruWs = rruWb['rru info']
print "Worksheet Title: ", rruWs.title

rruInfoDic = {}

for index in range( 2, rruWs.get_highest_row() + 1 ):
	siteId = findSiteId( str(rruWs.cell( row = index, column = 1).value) )
	sector = [
	rruWs.cell( row = index, column = 2 ).value, 
	rruWs.cell( row = index, column = 3 ).value, 
	rruWs.cell( row = index, column = 4 ).value 
	]
	rruInfoDic[siteId] = sector
print rruInfoDic

# umts cell
rncIdList = []
siteIdList = []
cellIdList = []
cellNameList = []
lacList = []
sacList = []
racList = []
uraList = []
pscList = []
uarfcnList = []
maxDlPowerList = []
pcpichList = []
tcellList = []
latitudeList = []
longitudeList = []
azimuthList = []
heightList = []

sectorEqmIdList = []
basebandEqmIdList = []
uplinkUarfcnList = []
bandIndicatorList = []
frequencyBandwidthList = []

logicalRncIdList = []
maxCoverageList = []
antennaOpeningList = []

electricalTiltList = []
mechanicalTiltList = []

# gsm cell
bscIdList = []
gCellSiteIdList = []
gCellNameList = []
gCellMccList = []
gCellMncList = []
gCellLacList = []
gCellIdList = []
gCellNccList = []
gCellBccList = []
gCellPowerList = []
gCellBcchList = []
gCellHsnList = []
gCellLatitudeIntList = []
gCellLatitudeDecimalList = []
gCellLongitudeIntList = []
gCellLongitudeDecimalList = []
gCellAzimuthList = []
gCellAltitudeIntList = []
gCellFreListList = []
gCellFreMaGroupListList = []
gCellPdchNumList = []
gCellSdcchNumList = []
gCellOspList = []
gCellNseiQueueList = []
gCellBvciBeginList = []
gCellTrxGroupList = []
gCellSectorEqmList = []
gCellLastCharacterList = []

umtsCellWsDict = {}
for index in range( 1, umtsCellWs.get_highest_column() + 1 ):
	key = umtsCellWs.cell(row = 1, column = index).value
	umtsCellWsDict[ str(key) ] = index

for index in range( 2, umtsCellWs.get_highest_row() + 1 ):
	rncIdList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['RNCID']).value )
	siteIdList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['SITEID']).value )
	cellIdList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['CELLID']).value )
	cellNameList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['CELLNAME']).value )
	lacList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['LAC']).value )
	sacList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['SAC']).value )
	racList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['RAC']).value )
	uraList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['URA']).value )
	uarfcnList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['UARFCN']).value )
	pscList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['PSC']).value )
	maxDlPowerList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['MaxDlPower']).value )
	pcpichList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['PCPICH']).value )
	tcellList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['TCELL']).value )
	latitudeList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Latitude']).value )
	longitudeList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Longitude']).value )
	azimuthList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Azimuth']).value )
	heightList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Height']).value )

	sectorEqmIdList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['SectorEqm ID']).value )
	basebandEqmIdList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['BasebandEqm ID']).value )
	uplinkUarfcnList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Uplink UARFCN']).value )
	bandIndicatorList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['BandIndicator']).value )
	frequencyBandwidthList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['FrequencyBandwidth']).value )

	logicalRncIdList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Logical RNCID']).value )
	maxCoverageList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Max Coverage']).value )
	antennaOpeningList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['Antenna Opening']).value )

	electricalTiltList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['electricalTilt']).value )
	mechanicalTiltList.append( umtsCellWs.cell(row = index, column = umtsCellWsDict['mechanicalTilt']).value )

# for parameter, col in gsmCellWsDict.items():
# 	print '(%s, %d)' % (parameter, col)
gsmCellWsDict = {}
for index in range( 1, gsmCellWs.get_highest_column() + 1 ):
	key = gsmCellWs.cell(row = 1, column = index).value
	gsmCellWsDict[ str(key) ] = index

for index in range( 2, gsmCellWs.get_highest_row() + 1 ):
	gCellSiteIdList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['SITEID']).value )
	bscIdList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['BSCID']).value )
	gCellNameList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['CELLNAME']).value )
	gCellIdList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['CELLID']).value )
	gCellMccList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['MCC']).value )
	gCellMncList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['MNC']).value )
	gCellLacList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['LAC']).value )
	gCellNccList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['NCC']).value )
	gCellBccList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['BCC']).value )
	gCellBcchList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['BCCH']).value )
	gCellFreListList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Non-main BCCH']).value )
	gCellPowerList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Power']).value )
	gCellPdchNumList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['PDCH Num']).value )
	gCellSdcchNumList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['SDCCH Num']).value )
	gCellHsnList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['HSN']).value )
	gCellFreMaGroupListList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Frequency List of MA Group']).value )
	gCellLatitudeIntList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Latitude Int']).value )
	gCellLatitudeDecimalList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Latitude Decimal']).value )
	gCellLongitudeIntList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Longitude Int']).value )
	gCellLongitudeDecimalList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Longitude Decimal']).value )
	gCellAzimuthList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Azimuth']).value )
	gCellAltitudeIntList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Altitude Int']).value )
	gCellOspList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['BSC OPC']).value )
	gCellNseiQueueList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['BSC NSEI queue']).value )
	gCellBvciBeginList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['BVCI begin']).value )
	gCellTrxGroupList.append( str(gsmCellWs.cell(row = index, column = gsmCellWsDict['TRXGROUP']).value) )
	gCellSectorEqmList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['SECTOREQM NUM']).value )
	gCellLastCharacterList.append( gsmCellWs.cell(row = index, column = gsmCellWsDict['Last Character']).value )

#----------------
#write umts cell Configuration into SummaryData Excel
#----------------

cellConfigurationWb = Workbook()
umtsCellConfigurationWs = cellConfigurationWb.create_sheet(0, 'UMTS Cell')
gsmCellConfigurationWs = cellConfigurationWb.create_sheet(1,'GSM Cell')
gsmTrxgroupConfigurationWs = cellConfigurationWb.create_sheet(2, 'GTRXGROUP')
gsmCellPtpbvcWs = cellConfigurationWb.create_sheet(3,'PTP BVC')

if currentSiteWsIndex == 0:
	umtsCellConfigurationContent = ['UMTS Local cell template','TRUE','NORMAL_CELL','0','FALSE',
	'REQUIRE','1','400','200','3G_CELL_new_Template_917', 'Manual','1','ACTIVATED','ACTIVATED','CELL_ANTENNA','DEG','MET','']
	print umtsCellConfigurationContent
	umtsCellConfigurationWs.append(
		{
		'A' : 'NodeB',
		'B' : 'UMTS Local Cell',
		'M' : 'UMTS Cell',
		'AH' : 'Cell HSDPA Parameters',
		'AK' : 'Cell HSUPA Parameters',
		'AM' : 'Cell SMLC Information'
		})
	umtsCellConfigurationWs.merge_cells('B1:L1')
	umtsCellConfigurationWs.merge_cells('M1:AG1')
	umtsCellConfigurationWs.merge_cells('AH1:AJ1')
	umtsCellConfigurationWs.merge_cells('AK1:AL1')
	umtsCellConfigurationWs.merge_cells('AM1:AW1')
	
	umtsCellConfigurationWs.append(
		{
		'A' : '*NodeB Name',
		'B' : '*Local Cell ID',
		'C' : 'Max Output Power',
		'D' : '*Local Cell Template',
		'E' : '*Sector Equipment ID',
		'F' : 'UL Baseband Equipment ID',
		'G' : 'DL Baseband Equipment ID',
		'H' : '*Uplink UARFCN',
		'I' : '*Downlink UARFCN',
		'J' : 'DL64QAM',
		'K' : 'Local Cell Type',
		'L' : 'Frequency Bandwidth(kHz)',
		'M' : 'Cell ID',
		'N' : '*Cell Name',
		'O' : '*CN Operator Group Index',
		'P' : '*Band Indicator',
		'Q' : 'UL Frequency Ind',
		'R' : 'Uplink UARFCN',
		'S' : '*Downlink UARFCN',
		'T' : '*DL Primary Scrambling Code',
		'U' : '*Location Area Code',
		'V' : '*Service Area Code',
		'W' : '*RAC Configuration Indication',
		'X' : 'Routing Area Code',
		'Y' : '*Service Priority Group Identity',
		'Z' : '*URA ID',
		'AA' : '*Time Offset[chip]',
		'AB' : 'Max Transmit Power of Cell',
		'AC' : 'PCPICH Transmit Power',
		'AD' : 'Max Transmit Power of PCPICH',
		'AE' : 'Min Transmit Power of PCPICH',
		'AF' : '*Local Cell ID',
		'AG' : '*Cell Template',
		'AH' : 'Allocate Code Mode',
		'AI' : 'Code Number for HS-PDSCH',
		'AJ' : 'Cell HSDPA state',
		'AK' : 'Code Number for E-AGCH',
		'AL' : 'Cell HSUPA state',
		'AM' : 'RNC ID',
		'AN' : 'Cell Location Setting Type',
		'AO' : 'Geo-coordinate Data Format',
		'AP' : 'Metrology',
		'AQ' : 'Cell Antenna Latitude',
		'AR' : 'Cell Antenna Longitude',
		'AS' : 'Cell Antenna Altitude',
		'AT' : 'Cell Antenna Max Coverage',
		'AU' : 'Cell Antenna Orientation',
		'AV' : 'Cell Antenna Opening',
		'AW' : 'Cell Average Altitude'
		})
	for index in range(1, len(siteIdList) + 1):
		umtsCellConfigurationWs.append(
			{
			'A' : str(siteIdList[index - 1]) + 'UMTS',
			'B' : cellIdList[index - 1],
			'C' : maxDlPowerList[index - 1],
			'D' : umtsCellConfigurationContent[0],
			'E' : sectorEqmIdList[index - 1],
			'F' : basebandEqmIdList[index - 1],
			'G' : basebandEqmIdList[index - 1],
			'H' : uplinkUarfcnList[index - 1],
			'I' : uarfcnList[index - 1],
			'J' : umtsCellConfigurationContent[1],
			'K' : umtsCellConfigurationContent[2],
			'L' : frequencyBandwidthList[index - 1],
			'M' : cellIdList[index - 1],
			'N' : cellNameList[index - 1],
			'O' : umtsCellConfigurationContent[3],
			'P' : bandIndicatorList[index - 1],
			'Q' : umtsCellConfigurationContent[4],
			'S' : uarfcnList[index - 1],
			'T' : pscList[index - 1],
			'U' : lacList[index - 1],
			'V' : sacList[index - 1],
			'W' : umtsCellConfigurationContent[5],
			'X' : racList[index - 1],
			'Y' : umtsCellConfigurationContent[6],
			'Z' : uraList[index -1],
			'AA' : tcellList[index - 1],
			'AB' : maxDlPowerList[index - 1],
			'AC' : pcpichList[index - 1],
			'AD' : umtsCellConfigurationContent[7],
			'AE' : umtsCellConfigurationContent[8],
			'AF' : cellIdList[index - 1],
			'AG' : umtsCellConfigurationContent[9],
			'AH' : umtsCellConfigurationContent[10],
			'AI' : umtsCellConfigurationContent[11],
			'AJ' : umtsCellConfigurationContent[12],
			'AL' : umtsCellConfigurationContent[13],
			'AM' : logicalRncIdList[index -1],
			'AN' : umtsCellConfigurationContent[14],
			'AO' : umtsCellConfigurationContent[15],
			'AP' : umtsCellConfigurationContent[16],
			'AQ' : latitudeList[index - 1],
			'AR' : longitudeList[index - 1],
			'AS' : heightList[index - 1],
			'AT' : maxCoverageList[index - 1],
			'AU' : azimuthList[index - 1],
			'AV' : antennaOpeningList[index - 1],
			'AW' : heightList[index - 1]
			})
elif currentSiteWsIndex == 1:
	umtsCellConfigurationContent = ['UMTS Local cell template','TRUE','NORMAL_CELL','0','FALSE',
	'REQUIRE','1','400','200','3G_CELL_new_Template_917', 'Manual','1','ACTIVATED','ACTIVATED','CELL_ANTENNA','DEG','MET','']
	print umtsCellConfigurationContent
	umtsCellConfigurationWs.append(
		{
		'A' : 'NodeB',
		'B' : 'UMTS Local Cell',
		'M' : 'UMTS Cell',
		'AH' : 'Cell HSDPA Parameters',
		'AK' : 'Cell HSUPA Parameters',
		'AM' : 'Cell SMLC Information'
		})
	umtsCellConfigurationWs.merge_cells('B1:L1')
	umtsCellConfigurationWs.merge_cells('M1:AG1')
	umtsCellConfigurationWs.merge_cells('AH1:AJ1')
	umtsCellConfigurationWs.merge_cells('AK1:AL1')
	umtsCellConfigurationWs.merge_cells('AM1:AW1')
	
	umtsCellConfigurationWs.append(
		{
		'A' : '*NodeB Name',
		'B' : '*Local Cell ID',
		'C' : 'Max Output Power',
		'D' : '*Local Cell Template',
		'E' : '*Sector Equipment ID',
		'F' : 'UL Baseband Equipment ID',
		'G' : 'DL Baseband Equipment ID',
		'H' : '*Uplink UARFCN',
		'I' : '*Downlink UARFCN',
		'J' : 'DL64QAM',
		'K' : 'Local Cell Type',
		'L' : 'Frequency Bandwidth(kHz)',
		'M' : 'Cell ID',
		'N' : '*Cell Name',
		'O' : '*CN Operator Group Index',
		'P' : '*Band Indicator',
		'Q' : 'UL Frequency Ind',
		'R' : 'Uplink UARFCN',
		'S' : '*Downlink UARFCN',
		'T' : '*DL Primary Scrambling Code',
		'U' : '*Location Area Code',
		'V' : '*Service Area Code',
		'W' : '*RAC Configuration Indication',
		'X' : 'Routing Area Code',
		'Y' : '*Service Priority Group Identity',
		'Z' : '*URA ID',
		'AA' : '*Time Offset[chip]',
		'AB' : 'Max Transmit Power of Cell',
		'AC' : 'PCPICH Transmit Power',
		'AD' : 'Max Transmit Power of PCPICH',
		'AE' : 'Min Transmit Power of PCPICH',
		'AF' : '*Local Cell ID',
		'AG' : '*Cell Template',
		'AH' : 'Allocate Code Mode',
		'AI' : 'Code Number for HS-PDSCH',
		'AJ' : 'Cell HSDPA state',
		'AK' : 'Code Number for E-AGCH',
		'AL' : 'Cell HSUPA state',
		'AM' : 'RNC ID',
		'AN' : 'Cell Location Setting Type',
		'AO' : 'Geo-coordinate Data Format',
		'AP' : 'Metrology',
		'AQ' : 'Cell Antenna Latitude',
		'AR' : 'Cell Antenna Longitude',
		'AS' : 'Cell Antenna Altitude',
		'AT' : 'Cell Antenna Max Coverage',
		'AU' : 'Cell Antenna Orientation',
		'AV' : 'Cell Antenna Opening',
		'AW' : 'Cell Average Altitude'
		})
	for index in range(1, len(siteIdList) + 1):
		umtsCellConfigurationWs.append(
			{
			'A' : str(siteIdList[index - 1]) + 'UMTS',
			'B' : cellIdList[index - 1],
			'C' : maxDlPowerList[index - 1],
			'D' : umtsCellConfigurationContent[0],
			'E' : sectorEqmIdList[index - 1],
			'F' : basebandEqmIdList[index - 1],
			'G' : basebandEqmIdList[index - 1],
			'H' : uplinkUarfcnList[index - 1],
			'I' : uarfcnList[index - 1],
			'J' : umtsCellConfigurationContent[1],
			'K' : umtsCellConfigurationContent[2],
			'L' : frequencyBandwidthList[index - 1],
			'M' : cellIdList[index - 1],
			'N' : cellNameList[index - 1],
			'O' : umtsCellConfigurationContent[3],
			'P' : bandIndicatorList[index - 1],
			'Q' : umtsCellConfigurationContent[4],
			'S' : uarfcnList[index - 1],
			'T' : pscList[index - 1],
			'U' : lacList[index - 1],
			'V' : sacList[index - 1],
			'W' : umtsCellConfigurationContent[5],
			'X' : racList[index - 1],
			'Y' : umtsCellConfigurationContent[6],
			'Z' : uraList[index -1],
			'AA' : tcellList[index - 1],
			'AB' : maxDlPowerList[index - 1],
			'AC' : pcpichList[index - 1],
			'AD' : umtsCellConfigurationContent[7],
			'AE' : umtsCellConfigurationContent[8],
			'AF' : cellIdList[index - 1],
			'AG' : umtsCellConfigurationContent[9],
			'AH' : umtsCellConfigurationContent[10],
			'AI' : umtsCellConfigurationContent[11],
			'AK' : umtsCellConfigurationContent[12]
			})
elif currentSiteWsIndex == 2:
	umtsCellConfigurationContent = ['UMTS Local cell template','TRUE','NORMAL_CELL','3G_CELL_new_Template_917','0','REQUIRE','FALSE',
	'200','400','1','CELL_ANTENNA','DEG','MET','ACTIVATED','Manual','1','ACTIVATED']
	print umtsCellConfigurationContent

	umtsCellConfigurationWs.append(
		{
		'A' : '*NodeB Name',
		'B' : '*Local Cell ID',
		'C' : '*Local Cell Template',
		'D' : '*Uplink UARFCN',
		'E' : '*Downlink UARFCN',
		'F' : '*Max Output Power',
		'G' : 'UL Baseband Equipment ID',
		'H' : 'DL Baseband Equipment ID',
		'I' : 'Sector Equipment ID',
		'J' : 'Azimuth',
		'K' : 'Electrical Tilt',
		'L' : 'Mechanical Tilt',
		'M' : 'DL64QAM',
		'N' : 'Local Cell Type',
		'O' : 'Frequency Bandwidth(kHz)',
		'P' : 'Cell ID',
		'Q' : '*Cell Name',
		'R' : '*Local Cell ID',
		'S' : '*Cell Template',
		'T' : '*CN Operator Group Index',
		'U' : '*Location Area Code',
		'V' : '*Service Area Code',
		'W' : '*RAC Configuration Indication',
		'X' : 'Routing Area Code',
		'Y' : '*Band Indicator',
		'Z' : 'UL Frequency Ind',
		'AA' : 'Uplink UARFCN',
		'AB' : '*Downlink UARFCN',
		'AC' : '*DL Primary Scrambling Code',
		'AD' : 'Min Transmit Power of PCPICH',
		'AE' : 'Max Transmit Power of PCPICH',
		'AF' : '*Service Priority Group Identity',
		'AG' : '*URA ID',
		'AH' : '*Time Offset[chip]',
		'AI' : 'Max Transmit Power of Cell',
		'AJ' : 'PCPICH Transmit Power',
		'AK' : 'RNC ID',
		'AL' : 'Cell Location Setting Type',
		'AM' : 'Geo-coordinate Data Format',
		'AN' : 'Metrology',
		'AO' : 'Cell Antenna Latitude',
		'AP' : 'Cell Antenna Longitude',
		'AQ' : 'Cell Antenna Altitude',
		'AR' : 'Cell Antenna Max Coverage',
		'AS' : 'Cell Antenna Orientation',
		'AT' : 'Cell Antenna Opening',
		'AU' : 'Cell Average Altitude',
		'AV' : 'Cell HSUPA state',
		'AW' : 'Allocate Code Mode',
		'AX' : 'Code Number for HS-PDSCH',
		'AY' : 'Cell HSDPA state'		
		})
	for index in range(1, len(siteIdList) + 1):
		umtsCellConfigurationWs.append(
			{
			'A' : str(siteIdList[index - 1]) + 'UMTS',
			'B' : cellIdList[index - 1],
			'C' : umtsCellConfigurationContent[0],
			'D' : uplinkUarfcnList[index - 1],
			'E' : uarfcnList[index - 1],
			'F' : maxDlPowerList[index - 1],
			'G' : basebandEqmIdList[index - 1],
			'H' : basebandEqmIdList[index - 1],
			'I' : sectorEqmIdList[index - 1],
			'J' : azimuthList[index - 1],
			'K' : electricalTiltList[index - 1],
			'L' : mechanicalTiltList[index - 1],
			'M' : umtsCellConfigurationContent[1],
			'N' : umtsCellConfigurationContent[2],
			'O' : frequencyBandwidthList[index - 1],
			'P' : cellIdList[index - 1],
			'Q' : cellNameList[index - 1],
			'R' : cellIdList[index - 1],
			'S' : umtsCellConfigurationContent[3],
			'T' : umtsCellConfigurationContent[4],
			'U' : lacList[index - 1],
			'V' : sacList[index - 1],
			'W' : umtsCellConfigurationContent[5],
			'X' : racList[index - 1],
			'Y' : bandIndicatorList[index - 1],
			'Z' : umtsCellConfigurationContent[6],
			'AB' : uarfcnList[index - 1],
			'AC' : pscList[index - 1],
			'AD' : umtsCellConfigurationContent[7],
			'AE' : umtsCellConfigurationContent[8],
			'AF' : umtsCellConfigurationContent[9],
			'AG' : uraList[index - 1],
			'AH' : tcellList[index - 1],
			'AI' : maxDlPowerList[index - 1],
			'AJ' : pcpichList[index - 1],
			'AK' : logicalRncIdList[index - 1],
			'AL' : umtsCellConfigurationContent[10],
			'AM' : umtsCellConfigurationContent[11],
			'AN' : umtsCellConfigurationContent[12],
			'AO' : latitudeList[index - 1],
			'AP' : longitudeList[index - 1],
			'AQ' : heightList[index - 1],
			'AR' : maxCoverageList[index - 1],
			'AS' : azimuthList[index - 1],
			'AT' : antennaOpeningList[index - 1],
			'AU' : heightList[index - 1],
			'AV' : umtsCellConfigurationContent[13],
			'AW' : umtsCellConfigurationContent[14],
			'AX' : umtsCellConfigurationContent[15],
			'AY' : umtsCellConfigurationContent[16]
			})

elif currentSiteWsIndex == 3:
	umtsCellConfigurationContent = ['UMTS Local cell template','TRUE','NORMAL_CELL','3G_CELL_new_Template_917','FALSE',
	'REQUIRE','1','400','200', '0','ACTIVATED','CELL_ANTENNA','DEG','MET','ACTIVATED']
	print umtsCellConfigurationContent
	
	umtsCellConfigurationWs.append(
		{
		'A' : '*NodeB Name',
		'B' : '*Local Cell ID',
		'C' : '*Local Cell Template',
		'D' : '*Downlink UARFCN',
		'E' : '*Uplink UARFCN',
		'F' : '*Max Output Power',
		'G' : 'Sector Equipment ID',
		'H' : 'Azimuth',
		'I' : 'Electrical Tilt',
		'J' : 'Mechanical Tilt',
		'K' : 'DL64QAM',
		'L' : 'Local Cell Type',
		'M' : 'Frequency Bandwidth(kHz)',
		'N' : 'Local Cell Radius',
		'O' : 'Local Cell Inner Handover Radius',
		'P' : 'UL Baseband Equipment ID',
		'Q' : 'DL Baseband Equipment ID',
		'R' : '*Local Cell ID',
		'S' : 'Cell ID',
		'T' : '*Cell Name',
		'U' : '*Cell Template',
		'V' : '*Band Indicator',
		'W' : 'UL Frequency Ind',
		'X' : '*Downlink UARFCN',
		'Y' : '*DL Primary Scrambling Code',
		'Z' : '*Location Area Code',
		'AA' : '*Service Area Code',
		'AB' : '*RAC Configuration Indication',
		'AC' : 'Routing Area Code',
		'AD' : '*Service Priority Group Identity',
		'AE' : '*URA ID',
		'AF' : 'Cell Oriented Cell Individual Offset',
		'AG' : '*Time Offset[chip]',
		'AH' : 'Max Transmit Power of Cell',
		'AI' : 'PCPICH Transmit Power',
		'AJ' : 'Max Transmit Power of PCPICH',
		'AK' : 'Min Transmit Power of PCPICH',
		'AL' : '*CN Operator Group Index',
		'AM' : 'Cell HSDPA state',
		'AN' : 'RNC ID',
		'AO' : 'Cell Location Setting Type',
		'AP' : 'Geo-coordinate Data Format',
		'AQ' : 'Metrology',
		'AR' : 'Cell Antenna Latitude',
		'AS' : 'Cell Antenna Longitude',
		'AT' : 'Cell Antenna Altitude',
		'AU' : 'Cell Antenna Max Coverage',
		'AV' : 'Cell Antenna Orientation',
		'AW' : 'Cell Antenna Opening',
		'AX' : 'Cell Average Altitude',
		'AY' : 'Cell HSUPA state'		
		})
	for index in range(1, len(siteIdList) + 1):
		umtsCellConfigurationWs.append(
			{
			'A' : str(siteIdList[index - 1]) + 'UMTS',
			'B' : cellIdList[index - 1],
			'C' : umtsCellConfigurationContent[0],
			'D' : uarfcnList[index - 1],
			'E' : uplinkUarfcnList[index - 1],
			'F' : maxDlPowerList[index - 1],
			'G' : sectorEqmIdList[index - 1],
			'H' : azimuthList[index - 1],
			'I' : electricalTiltList[index - 1],
			'J' : mechanicalTiltList[index - 1],
			'K' : umtsCellConfigurationContent[1],
			'L' : umtsCellConfigurationContent[2],
			'M' : frequencyBandwidthList[index - 1],		
			'P' : basebandEqmIdList[index - 1],
			'Q' : basebandEqmIdList[index - 1],
			'R' : cellIdList[index - 1],
			'S' : cellIdList[index - 1],
			'T' : cellNameList[index - 1],
			'U' : umtsCellConfigurationContent[3],
			'V' : bandIndicatorList[index - 1],
			'W' : umtsCellConfigurationContent[4],
			'X' : uarfcnList[index - 1],
			'Y' : pscList[index - 1],
			'Z' : lacList[index - 1],
			'AA' : sacList[index - 1],
			'AB' : umtsCellConfigurationContent[5],
			'AC' : racList[index - 1],
			'AD' : umtsCellConfigurationContent[6],
			'AE' : uraList[index -1],
			'AG' : tcellList[index - 1],
			'AH' : maxDlPowerList[index - 1],
			'AI' : pcpichList[index - 1],
			'AJ' : umtsCellConfigurationContent[7],
			'AK' : umtsCellConfigurationContent[8],
			'AL' : umtsCellConfigurationContent[9],
			'AM' : umtsCellConfigurationContent[10],
			'AN' : logicalRncIdList[index - 1],
			'AO' : umtsCellConfigurationContent[11],
			'AP' : umtsCellConfigurationContent[12],
			'AQ' : umtsCellConfigurationContent[13],
			'AR' : latitudeList[index - 1],
			'AS' : longitudeList[index - 1],
			'AT' : heightList[index - 1],
			'AU' : maxCoverageList[index - 1],
			'AV' : azimuthList[index - 1],
			'AW' : antennaOpeningList[index - 1],
			'AX' : heightList[index - 1],
			'AY' : umtsCellConfigurationContent[14]
			})	
elif currentSiteWsIndex == 4:
	umtsCellConfigurationContent = ['0','1','0','FALSE','ACTIVATED']
	print umtsCellConfigurationContent	

	umtsCellConfigurationWs.append(
		{
		'A' : 'BSC Name',
		'B' : 'Logical RNC ID',
		'C' : 'NodeB Name',
		'D' : 'Cell ID',
		'E' : 'Cell Name',
		'F' : 'Cn Operator Group Index',
		'G' : 'Band Indicator',
		'H' : 'Uplink UARFCN',
		'I' : 'Downlink UARFCN',
		'J' : 'DL Primary Scrambling Code',
		'K' : 'Location Area Code',
		'L' : 'Service Area Code',
		'M' : 'Routing Area Code',
		'N' : 'Service Priority Group Identity',
		'O' : 'URA ID',
		'P' : 'Cell Oriented Cell Individual Offset',
		'Q' : 'Time Offset',
		'R' : 'Max Transmit Power of Cell',
		'S' : 'PCPICH Transmit Power',
		'T' : 'Local Cell ID',
		'U' : 'Validation indication',
		'V' : 'Heterogeneous Cell Flag',
		'W' : 'Cell Template Name'		
		})
	for index in range(1, len(siteIdList) + 1):
		umtsCellConfigurationWs.append(
			{
			'A' : rncIdList[index - 1],
			'B' : logicalRncIdList[index - 1],
			'C' : str(siteIdList[index - 1]) + 'UMTS',
			'D' : cellIdList[index - 1],
			'E' : cellNameList[index - 1],
			'F' : umtsCellConfigurationContent[0],
			'G' : bandIndicatorList[index - 1],
			'I' : uarfcnList[index - 1],
			'J' : pscList[index - 1],
			'K' : lacList[index - 1],
			'L' : sacList[index - 1],
			'M' : racList[index - 1],
			'N' : umtsCellConfigurationContent[1],
			'O' : uraList[index - 1],
			'P' : umtsCellConfigurationContent[2],
			'Q' : tcellList[index - 1],
			'R' : maxDlPowerList[index - 1],
			'S' : pcpichList[index - 1],
			'T' : cellIdList[index - 1],
			'U' : umtsCellConfigurationContent[4],
			'V' : umtsCellConfigurationContent[3]
			})
elif currentSiteWsIndex == 5:
	umtsCellConfigurationContent = ['0','1','0','FALSE','ACTIVATED']
	print umtsCellConfigurationContent	

	umtsCellConfigurationWs.append(
		{
		'A' : 'BSC Name',
		'B' : 'Logical RNC ID',
		'C' : 'Cell ID',
		'D' : 'Cell Name',
		'E' : 'Max Transmit Power of Cell',
		'F' : 'Band Indicator',
		'G' : 'Cn Operator Group Index',
		'H' : 'Uplink UARFCN',
		'I' : 'Downlink UARFCN',
		'J' : 'Time Offset',
		'K' : 'DL Primary Scrambling Code',
		'L' : 'Service Priority Group Identity',
		'M' : 'NodeB Name',
		'N' : 'Local Cell ID',
		'O' : 'Location Area Code',
		'P' : 'Service Area Code',
		'Q' : 'Routing Area Code',
		'R' : 'Cell Oriented Cell Individual Offset',
		'S' : 'Heterogeneous Cell Flag',
		'T' : 'Validation indication',
		'U' : 'URA ID',
		'V' : 'PCPICH Transmit Power',
		'W' : 'Cell Template Name'	
		})
	for index in range(1, len(siteIdList) + 1):
		umtsCellConfigurationWs.append(
			{
			'A' : rncIdList[index - 1],
			'B' : logicalRncIdList[index - 1],
			'M' : str(siteIdList[index - 1]) + 'UMTS',
			'C' : cellIdList[index - 1],
			'D' : cellNameList[index - 1],
			'G' : umtsCellConfigurationContent[0],
			'F' : bandIndicatorList[index - 1],
			'I' : uarfcnList[index - 1],
			'K' : pscList[index - 1],
			'O' : lacList[index - 1],
			'P' : sacList[index - 1],
			'Q' : racList[index - 1],
			'L' : umtsCellConfigurationContent[1],
			'U' : uraList[index - 1],
			'R' : umtsCellConfigurationContent[2],
			'J' : tcellList[index - 1],
			'E' : maxDlPowerList[index - 1],
			'V' : pcpichList[index - 1],
			'N' : cellIdList[index - 1]，
			'S' : umtsCellConfigurationContent[3],
			'T' : umtsCellConfigurationContent[4]
			})

if currentSiteWsIndex == 0:
	gsmCellConfigurationContent = ['GBTS_Cell','2G Cell template_917','PCS1900','Normal_cell','ON',
	'Extra','OFF','NO','NO','YES','40','0','MBCCH','RF_FH','NONE','SupportAsInnPcu','YES','NO','YES','NO','10',
	'ALWAYSNORMALPAGING','SUPPORT','Degree','North_latitude','West_Longitude']
	print gsmCellConfigurationContent
	gsmCellConfigurationWs.append(
		{
		'A' : 'eGBTS',
		'B' : 'GSM Local Cell',
		'D' : 'GSM CELL',
		'Q' : 'Call Control AMR Parameters of Cell',
		'V' : 'Relation between Cell and OSP',
		'W' : 'TRX Info',
		'AJ' : 'Concentric Attributes of TRX',
		'AK' : 'Basic GPRS Attributes of 2G Cell',
		'AS' : 'LCS Parameters of Cell'			
		})
	gsmCellConfigurationWs.merge_cells('B1:C1')
	gsmCellConfigurationWs.merge_cells('D1:P1')
	gsmCellConfigurationWs.merge_cells('Q1:U1')
	gsmCellConfigurationWs.merge_cells('W1:AI1')
	gsmCellConfigurationWs.merge_cells('AK1:AR1')
	gsmCellConfigurationWs.merge_cells('AS1:BA1')
	
	gsmCellConfigurationWs.append(
		{
		'A' : '*BTS Name',
		'B' : '*LoCellID',
		'C' : '*Local Cell Template',
		'D' : '*GSM Cell Name',
		'E' : '*GSM Cell Template Name',
		'F' : '*Cell Type',
		'G' : '*MCC',
		'H' : '*MNC',
		'I' : '*LAC',
		'J' : '*CI',
		'K' : 'NCC',
		'L' : 'BCC',
		'M' : 'Cell IUO Type',
		'N' : 'Enhanced Concentric Allowed',
		'O' : 'BCCH IUO of Double Frequency Cell',
		'P' : 'Start Flex MAIO Switch',
		'Q' : 'Auto Adjust UL TH and Hysteresis [H]',
		'R' : 'AMR Rate Control Switch',
		'S' : 'Auto Adjust UL TH and Hysteresis [F]',
		'T' : 'AMR Uplink Adaptive Threshold Allowed',
		'U' : 'Uplink Threshold Adjust Factor',
		'V' : 'Cell OSP Map',
		'W' : '*Frequency of BCCH',
		'X' : 'Non-Main BCCH Frequency List (Separated by ",")',
		'Y' : 'Power Level',
		'Z' : 'eGBTS Power Type(0.1dBm)',
		'AA' : 'BCCH Type',
		'AB' : 'Number of PDCH',
		'AC' : 'Number of SDCCH',
		'AD' : 'Number of BCH',
		'AE' : 'Support CBCH',
		'AF' : 'Hop Type',
		'AG' : 'Hopping sequence number',
		'AH' : 'Frequency List of MA Group',
		'AI' : 'TRX Group ID',
		'AJ' : 'Concentric Attribute',
		'AK' : 'GPRS Support',
		'AL' : 'EDGE',
		'AM' : 'PACKET SI',
		'AN' : 'Support NACC',
		'AO' : 'Support eNACC',
		'AP' : 'Routing Area',
		'AQ' : 'PS Paging Control',
		'AR' : 'Support EDA',
		'AS' : 'Latitude and Longitude Input Mode',
		'AT' : 'NS Latitude',
		'AU' : 'Latitude Int Part',
		'AV' : 'Latitude Decimal Part',
		'AW' : 'WE Longitude',
		'AX' : 'Longitude Int Part',
		'AY' : 'Longitude Decimal Part',
		'AZ' : 'Antenna Azimuth Angle',
		'BA' : 'Antenna Altitude Int Part'
		})
	
	for index in range(1, len(gCellSiteIdList) + 1):
		gsmCellConfigurationWs.append(
			{
			'A' : gCellSiteIdList[index - 1] + 'GSM',
			'B' : gCellIdList[index - 1],
			'C' : gsmCellConfigurationContent[0],
			'D' : gCellNameList[index - 1],
			'E' : gsmCellConfigurationContent[1],
			'F' : gsmCellConfigurationContent[2],
			'G' : gCellMccList[index - 1],
			'H' : gCellMncList[index - 1],
			'I' : gCellLacList[index - 1],
			'J' : gCellIdList[index - 1],
			'K' : gCellNccList[index - 1],
			'L' : gCellBccList[index - 1],
			'M' : gsmCellConfigurationContent[3],
			'N' : gsmCellConfigurationContent[4],
			'O' : gsmCellConfigurationContent[5],
			'P' : gsmCellConfigurationContent[6],
			'Q' : gsmCellConfigurationContent[7],
			'S' : gsmCellConfigurationContent[8],
			'T' : gsmCellConfigurationContent[9],
			'U' : gsmCellConfigurationContent[10],
			'V' : gCellOspList[index - 1],
			'W' : gCellBcchList[index - 1],
			'X' : gCellFreListList[index - 1],
			'Y' : gsmCellConfigurationContent[11],
			'Z' : gCellPowerList[index - 1],
			'AA' : gsmCellConfigurationContent[12],
			'AB' : gCellPdchNumList[index - 1],
			'AC' : gCellSdcchNumList[index - 1],
			'AF' : gsmCellConfigurationContent[13],
			'AG' : gCellHsnList[index - 1],
			'AH' : gCellFreMaGroupListList[index - 1],
			'AI' : gCellTrxGroupList[index - 1],
			'AJ' : gsmCellConfigurationContent[14],
			'AK' : gsmCellConfigurationContent[15],
			'AL' : gsmCellConfigurationContent[16],
			'AM' : gsmCellConfigurationContent[17],
			'AN' : gsmCellConfigurationContent[18],
			'AO' : gsmCellConfigurationContent[19],
			'AP' : gsmCellConfigurationContent[20],
			'AQ' : gsmCellConfigurationContent[21],
			'AR' : gsmCellConfigurationContent[22],
			'AS' : gsmCellConfigurationContent[23],
			'AT' : gsmCellConfigurationContent[24],
			'AU' : gCellLatitudeIntList[index - 1],
			'AV' : gCellLatitudeDecimalList[index - 1],
			'AW' : gsmCellConfigurationContent[25],
			'AX' : gCellLongitudeIntList[index - 1],
			'AY' : gCellLongitudeDecimalList[index - 1],
			'AZ' : gCellAzimuthList[index - 1],
			'BA' : gCellAltitudeIntList[index - 1]
			})
elif currentSiteWsIndex == 2:
	gsmCellConfigurationContent = ['GBTS_Cell','2G Cell template_917','PCS1900','Normal_cell','ON',
	'Extra','OFF','NO','NO','YES','40','0','MBCCH','RF_FH','NONE','SupportAsInnPcu','YES','North_latitude','Degree','West_Longitude']
	print gsmCellConfigurationContent
	gsmCellConfigurationWs.append(
		{
		'A' : '*BTS Name',
		'B' : '*LoCellID',
		'C' : '*Local Cell Template',
		'D' : '*GSM Cell Name',
		'E' : '*GSM Cell Template Name',
		'F' : '*Cell Type',
		'G' : '*MCC',
		'H' : '*MNC',
		'I' : '*LAC',
		'J' : '*CI',
		'K' : 'NCC',
		'L' : 'BCC',
		'M' : 'Cell IUO Type',
		'N' : 'Enhanced Concentric Allowed',
		'O' : 'BCCH IUO of Double Frequency Cell',
		'P' : 'Start Flex MAIO Switch',
		'Q' : 'Auto Adjust UL TH and Hysteresis [H]',
		'R' : 'Auto Adjust UL TH and Hysteresis [F]',
		'S' : 'AMR Rate Control Switch',
		'T' : 'AMR Uplink Adaptive Threshold Allowed',
		'U' : 'Uplink Threshold Adjust Factor',
		'V' : 'OSP Code',
		'W' : 'Cell OSP Map',
		'X' : 'TRX Group ID',
		'Y' : '*Frequency of BCCH',
		'Z' : 'Non-Main BCCH Frequency List (Separated by ",")',
		'AA' : 'Power Level',
		'AB' : 'eGBTS Power Type(0.1dBm)',
		'AC' : 'BCCH Type',
		'AD' : 'Number of PDCH',
		'AE' : 'Number of SDCCH',
		'AF' : 'Hop Type',
		'AG' : 'Hopping sequence number',
		'AH' : 'Frequency List of MA Group',
		'AI' : 'Concentric Attribute',
		'AJ' : 'GPRS Support',
		'AK' : 'EDGE',
		'AL' : 'Longitude Decimal Part',
		'AM' : 'Latitude Int Part',
		'AN' : 'NS Latitude',
		'AO' : 'Latitude and Longitude Input Mode',
		'AP' : 'Latitude Decimal Part',
		'AQ' : 'Antenna Altitude Int Part',
		'AR' : 'Antenna Altitude Decimal Part',
		'AS' : 'Antenna Azimuth Angle',
		'AT' : 'WE Longitude',
		'AU' : 'Longitude Int Part'
		})
	for index in range(1, len(gCellSiteIdList) + 1):
		gsmCellConfigurationWs.append(
			{
			'A' : gCellSiteIdList[index - 1] + 'GSM',
			'B' : gCellIdList[index - 1],
			'C' : gsmCellConfigurationContent[0],
			'D' : gCellNameList[index - 1],
			'E' : gsmCellConfigurationContent[1],
			'F' : gsmCellConfigurationContent[2],
			'G' : gCellMccList[index - 1],
			'H' : gCellMncList[index - 1],
			'I' : gCellLacList[index - 1],
			'J' : gCellIdList[index - 1],
			'K' : gCellNccList[index - 1],
			'L' : gCellBccList[index - 1],
			'M' : gsmCellConfigurationContent[3],
			'N' : gsmCellConfigurationContent[4],
			'O' : gsmCellConfigurationContent[5],
			'P' : gsmCellConfigurationContent[6],
			'Q' : gsmCellConfigurationContent[7],
			'R' : gsmCellConfigurationContent[8],
			'T' : gsmCellConfigurationContent[9],
			'U' : gsmCellConfigurationContent[10],
			'W' : gCellOspList[index - 1],
			'X' : gCellTrxGroupList[index - 1],
			'Y' : gCellBcchList[index - 1],
			'Z' : gCellFreListList[index - 1],
			'AA' : gsmCellConfigurationContent[11],
			'AB' : gCellPowerList[index - 1],
			'AC' : gsmCellConfigurationContent[12],
			'AD' : gCellPdchNumList[index - 1],
			'AE' : gCellSdcchNumList[index - 1],
			'AF' : gsmCellConfigurationContent[13],
			'AG' : gCellHsnList[index - 1],
			'AH' : gCellFreMaGroupListList[index - 1],
			'AI' : gsmCellConfigurationContent[14],
			'AJ' : gsmCellConfigurationContent[15],
			'AK' : gsmCellConfigurationContent[16],
			'AL' : gCellLongitudeDecimalList[index - 1],
			'AM' : gCellLatitudeIntList[index - 1],
			'AN' : gsmCellConfigurationContent[17],
			'AO' : gsmCellConfigurationContent[18],
			'AP' : gCellLatitudeDecimalList[index - 1],
			'AQ' : gCellAltitudeIntList[index - 1],
			'AS' : gCellAzimuthList[index - 1],
			'AT' : gsmCellConfigurationContent[19],
			'AU' : gCellLongitudeIntList[index - 1]
			})

if currentSiteWsIndex == 0 or currentSiteWsIndex == 2:
	gsmTrxgroupConfigurationWs.append(
		{
		'A' : 'eGBTS',
		'B' : 'GBTS TRX Group'	
		})
	gsmTrxgroupConfigurationWs.merge_cells('B1:I1')
	gsmTrxgroupConfigurationWs.append(
		{
		'A' : '*BTS Name',
		'B' : '*TRX Group ID',
		'C' : '*Local Cell ID',
		'D' : 'Sending Mode',
		'E' : 'Receiving Mode',
		'F' : 'Working Mode',
		'G' : 'User Label',
		'H' : '*Sector Equipment ID',
		'I' : 'Maximum Power(0.1dBm)'
		})
	gsmTrxgroupCommonContent = ['SINGLESND','MAINDIVERSITY','INDEPENDENT','65535']
	gsmTrxgroupContent = [
	['1','10'],
	['2','11'],
	['3','12'],
	['4','13'],
	['5','14'],
	['6','15'],
	['?','?'],
	['?','?'],
	['?','?'],
	['?','?'],
	['?','?'],
	['?','?'],
	['?','?'],
	['?','?']
	]
	for index in range( 1, len(gCellSiteIdList) + 1 ):
		if gCellLastCharacterList[index - 1]:
			lastCharacterNum = int(gCellLastCharacterList[index - 1])
			sectorScenario = rruInfoDic[ gCellSiteIdList[index - 1] ][ lastCharacterNum - 1 ]
		else:
			lastCharacterNum = 7
			sectorScenario = 0
		if( sectorScenario == 13 or sectorScenario == 14 ):
			gsmTrxgroupConfigurationWs.append(
				{
				'A' : gCellSiteIdList[index - 1] + 'GSM',
				'B' : gsmTrxgroupContent[ lastCharacterNum - 1 ][0],
				'C' : gCellIdList[index - 1],
				'D' : gsmTrxgroupCommonContent[0],
				'E' : gsmTrxgroupCommonContent[1],
				'F' : gsmTrxgroupCommonContent[2],
				'H' : gsmTrxgroupContent[ lastCharacterNum - 1 ][1],
				'I' : gsmTrxgroupCommonContent[3]
				})
			if( gCellSectorEqmList[index - 1] and int(gCellSectorEqmList[index - 1]) == 2 ):
				gsmTrxgroupConfigurationWs.append(
					{
					'A' : gCellSiteIdList[index - 1] + 'GSM',
					'B' : gsmTrxgroupContent[ lastCharacterNum + 2 ][0],
					'C' : gCellIdList[index - 1],
					'D' : gsmTrxgroupCommonContent[0],
					'E' : gsmTrxgroupCommonContent[1],
					'F' : gsmTrxgroupCommonContent[2],
					'H' : gsmTrxgroupContent[ lastCharacterNum + 2 ][1],
					'I' : gsmTrxgroupCommonContent[3]
					})
		elif( sectorScenario == 16 or sectorScenario == 29 ):
			gsmTrxgroupConfigurationWs.append(
				{
				'A' : gCellSiteIdList[index - 1] + 'GSM',
				'B' : gsmTrxgroupContent[ lastCharacterNum - 1 ][0],
				'C' : gCellIdList[index - 1],
				'D' : gsmTrxgroupCommonContent[0],
				'E' : gsmTrxgroupCommonContent[1],
				'F' : gsmTrxgroupCommonContent[2],
				'H' : int(gsmTrxgroupContent[ lastCharacterNum - 1 ][1]) - 3,
				'I' : gsmTrxgroupCommonContent[3]
				})
			if( gCellSectorEqmList[index - 1] and int(gCellSectorEqmList[index - 1]) == 2 ):
				gsmTrxgroupConfigurationWs.append(
					{
					'A' : gCellSiteIdList[index - 1] + 'GSM',
					'B' : gsmTrxgroupContent[ lastCharacterNum + 2 ][0],
					'C' : gCellIdList[index - 1],
					'D' : gsmTrxgroupCommonContent[0],
					'E' : gsmTrxgroupCommonContent[1],
					'F' : gsmTrxgroupCommonContent[2],
					'H' : int(gsmTrxgroupContent[ lastCharacterNum + 2 ][1]) - 3,
					'I' : gsmTrxgroupCommonContent[3]
					})	
		else:			
			gsmTrxgroupConfigurationWs.append(
				{
				'A' : gCellSiteIdList[index - 1] + 'GSM',
				'B' : gsmTrxgroupContent[ lastCharacterNum - 1 ][0],
				'C' : gCellIdList[index - 1],
				'D' : gsmTrxgroupCommonContent[0],
				'E' : gsmTrxgroupCommonContent[1],
				'F' : gsmTrxgroupCommonContent[2],
				'H' : '-999',
				'I' : gsmTrxgroupCommonContent[3]
				})	
	gsmCellPtpbvcWs.append(
		{
		'A' : 'GSM CELL',
		'B' : 'PTP BVC'
		})
	gsmCellPtpbvcWs.merge_cells('B1:C1')
	gsmCellPtpbvcWs.append(
		{
		'A' : '*GSM Cell Name',
		'B' : 'NSE Identifier',
		'C' : 'PTP BVC Identifier'
		})
	for index in range( 1, len(gCellSiteIdList) + 1 ):
		nseiList = str(gCellNseiQueueList[index - 1])
		nseiAfterSplitList = nseiList.split(',')
		for i in range(0,len(nseiAfterSplitList)):
			gsmCellPtpbvcWs.append(
				{
				'A' : gCellNameList[index - 1],
				'B' : nseiAfterSplitList[i],
				'C' : gCellBvciBeginList[index - 1] + i
				})





folder = time.strftime(r"cell_%Y-%m-%d_%H-%M-%S",time.localtime())
os.makedirs(r'%s/%s'%(os.getcwd(),folder))

if currentSiteWsIndex == 0:
    folder = folder + '\\GUL_'
elif currentSiteWsIndex == 1:
    folder = folder + '\\UL_'
elif currentSiteWsIndex == 2:
    folder = folder + '\\GU_'
elif currentSiteWsIndex == 3:
    folder = folder + '\\UO_'
elif currentSiteWsIndex == 4:
	folder = folder + '\\add UCELL in RNP(for offline-CME)_'
elif currentSiteWsIndex == 5:
	folder = folder + '\\add UCELL in RNP(for online-CME)_'

cellConfigurationWb.save( folder + 'Cell Configuration.xlsx' )
